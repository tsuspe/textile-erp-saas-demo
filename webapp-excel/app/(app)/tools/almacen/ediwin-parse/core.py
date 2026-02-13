# app/(app)/tools/almacen/ediwin-parse/core.py
# (SIN CAMBIOS: ya replica el parser del original + export ZIP + carpetas + split TXT)

# tools/ediwin-parser/core.py
# Core sin Streamlit. Genera salida completa (PDF/Excel/CSV + carpetas por modelo + TXT repartidos)
# y empaqueta en ZIP.

import os
import re
import zipfile
from collections import defaultdict
from io import BytesIO
from pathlib import Path
from typing import Dict, Optional, Set, Tuple

import pandas as pd
import pdfplumber
from openpyxl.styles import Border, Font, PatternFill, Side
from pypdf import PdfReader, PdfWriter

# ===================== CONSTANTES =====================

TALLAS = [
    "XXS",
    "XS",
    "S",
    "M",
    "L",
    "XL",
    "XXL",
    "34",
    "36",
    "38",
    "40",
    "42",
    "44",
    "46",
    "48",
]

SIZE_CODE_MAP = {
    "001": "XXS",
    "002": "XS",
    "003": "S",
    "004": "M",
    "005": "L",
    "006": "XL",
    "007": "XXL",
    "034": "34",
    "036": "36",
    "038": "38",
    "040": "40",
    "042": "42",
    "044": "44",
    "046": "46",
    "048": "48",
}


# ===================== UTILIDADES =====================


def safe_fs_name(name: str) -> str:
    name = name.replace("/", "-").replace("\\", "-")
    name = re.sub(r"[^\w\-]+", "_", name)
    return name.strip("_") or "SIN_NOMBRE"


def get_active_sizes(df: pd.DataFrame):
    activas = []
    for t in TALLAS:
        if t in df.columns:
            total = pd.to_numeric(df[t], errors="coerce").fillna(0).sum()
            if total > 0:
                activas.append(t)
    return activas


from openpyxl.styles import PatternFill

PALETTE = [
    "FDE2E4",
    "BEE1E6",
    "E2F0CB",
    "FFDFBA",
    "D0BDF4",
    "C7F9CC",
    "FFCCD5",
    "F1F0FF",
    "E5F4E3",
    "FFE5B4",
    "E0BBFF",
    "CAFFBF",
]


def _hex_fill(rgb_hex_no_hash: str):
    return PatternFill(
        start_color=rgb_hex_no_hash, end_color=rgb_hex_no_hash, fill_type="solid"
    )


def apply_model_row_colors(ws, modelo_col_name: str = "MODELO"):
    """
    Pinta filas completas según el valor de la columna MODELO.
    Replica el “style_by_model” del Streamlit original, pero en Excel real.
    """
    # localizar índice de columna MODELO en la fila 1 (headers)
    headers = [c.value for c in ws[1]]
    try:
        model_idx = headers.index(modelo_col_name) + 1  # 1-based
    except ValueError:
        return  # si no existe, no hacemos nada

    # asignar color por modelo
    model_colors = {}
    color_i = 0

    # recorrer filas desde la 2 (datos)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        model_val = row[model_idx - 1].value
        if model_val is None or str(model_val).strip() == "":
            continue

        key = str(model_val).strip()
        if key.upper() == "TOTAL":
            continue

        if key not in model_colors:
            model_colors[key] = PALETTE[color_i % len(PALETTE)]
            color_i += 1

        fill = _hex_fill(model_colors[key])

        for cell in row:
            # no sobrepintar celdas vacías si no quieres: aquí pintamos toda la fila como el original
            cell.fill = fill


def style_workbook_with_borders_and_headers(workbook):
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="FFFF00")
    header_font = Font(bold=True)

    total_fill = PatternFill("solid", fgColor="FFFF00")
    total_font = Font(bold=True)

    for ws in workbook.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column

        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font

        for row_idx in range(2, max_row + 1):
            first_cell = ws.cell(row=row_idx, column=1)
            if str(first_cell.value).strip().upper() == "TOTAL":
                for col_idx in range(1, max_col + 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.fill = total_fill
                    c.font = total_font


def zip_dir(src_dir: str, out_zip_path: str):
    src = Path(src_dir)
    out = Path(out_zip_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for p in src.rglob("*"):
            if p.is_file():
                z.write(p, arcname=str(p.relative_to(src)))


# ===================== PARSER EUROFIEL =====================


def split_orders(full_text: str):
    matches = list(re.finditer(r"Nº Pedido\s*:", full_text))
    chunks = []

    for i, m in enumerate(matches):
        start = m.start()
        prev_nl = full_text.rfind("\n", 0, start)
        if prev_nl != -1:
            prev_prev_nl = full_text.rfind("\n", 0, prev_nl)
            order_start = prev_prev_nl + 1 if prev_prev_nl != -1 else 0
        else:
            order_start = 0

        end = matches[i + 1].start() if i + 1 < len(matches) else len(full_text)
        chunk = full_text[order_start:end].strip()
        chunks.append(chunk)

    return chunks


def parse_detail_line_eurofiel(line: str):
    parts = line.split()
    if len(parts) < 8:
        return None
    if not parts[0].isdigit():
        return None
    if not re.fullmatch(r"\d{13}", parts[1]):
        return None

    cli_idx = None
    for i in range(2, len(parts)):
        if re.fullmatch(r"\d+/\d+/\d+", parts[i]):
            cli_idx = i
            break
    if cli_idx is None or cli_idx + 4 >= len(parts):
        return None

    cod_prov_full = " ".join(parts[2:cli_idx])
    cod_cli_full = parts[cli_idx]

    try:
        qty = int(parts[cli_idx + 1])
    except Exception:
        qty = 0

    p_neto = parts[cli_idx + 3]
    precio = p_neto.replace(",", ".")

    modelo = re.sub(r"/[^/]+$", "", cod_prov_full)
    patron = re.sub(r"/[^/]+$", "", cod_cli_full)
    talla = cod_prov_full.split("/")[-1].strip().upper()

    return modelo, patron, talla, qty, precio


def parse_order_eurofiel(order_text: str):
    lines = [ln for ln in order_text.splitlines() if ln.strip()]
    first_line = lines[0].strip() if lines else ""
    tipo = first_line

    def search(pattern: str):
        m = re.search(pattern, order_text)
        return m.group(1).strip() if m else ""

    pedido = search(r"Nº Pedido\s*:\s*(\S+)")
    fecha_entrega = search(r"Fecha Entrega\s*:\s*(\d{2}/\d{2}/\d{4})")

    pais = ""
    m_pais = re.search(r"País:\s*\([^)]*\)\s*([A-ZÁÉÍÓÚÜÑ ]+)", order_text)
    if m_pais:
        pais = m_pais.group(1).strip()

    descripcion = search(r"Descripción:\s*(.+)")
    total_unidades = search(r"Total Unidades\s+(\d+)")

    modelo = ""
    patron = ""
    precio = ""

    tallas_count = defaultdict(int)
    primer_detalle = True

    for ln in lines:
        parsed = parse_detail_line_eurofiel(ln)
        if parsed:
            m, p, talla, qty, prec = parsed
            if primer_detalle:
                modelo, patron, precio = m, p, prec
                primer_detalle = False
            tallas_count[talla.upper()] += qty

    row = {
        "TIPO": tipo,
        "PEDIDO": pedido,
        "FECHA_ENTREGA": fecha_entrega,
        "PAIS": pais,
        "DESCRIPCION": descripcion,
        "MODELO": modelo,
        "PATRON": patron,
        "PRECIO": precio,
        "TOTAL_UNIDADES": total_unidades,
    }

    for t in TALLAS:
        row[t] = tallas_count.get(t, 0)

    return row


def parse_pdf_eurofiel_bytes(pdf_bytes: bytes):
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    orders = split_orders(full_text)
    rows = [parse_order_eurofiel(o) for o in orders]
    rows = [r for r in rows if r.get("PEDIDO")]
    return pd.DataFrame(rows)


# (… el resto del archivo core.py va igual que el que me pasaste; no lo toco.)


# ===================== PARSER ECI =====================


def parse_page_eci(text: str):
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    def search(pattern: str):
        m = re.search(pattern, text)
        return m.group(1).strip() if m else ""

    tipo = ""
    for ln in lines:
        low = ln.lower()
        if low in (
            "pedido",
            "reposicion",
            "reposición",
            "anulacion pedido",
            "anulación pedido",
        ):
            tipo = ln.upper()
            break

    n_pedido = search(r"Nº Pedido\s+(\d+)")
    departamento = search(r"Dpto\. venta\s+(\d+)")
    fecha_entrega = search(r"Fecha Entrega\s+(\d{2}/\d{2}/\d{4})")

    suc_entrega = search(r"Sucursal Destino que Pide\s+([0-9 ]+)\s+[A-ZÁÉÍÓÚÜÑ]")
    if not suc_entrega:
        suc_entrega = search(r"Sucursal de Entrega\s+([0-9 ]+)\s+[A-ZÁÉÍÓÚÜÑ]")

    rows = []

    for i, ln in enumerate(lines):
        if not re.match(r"^\d+\s+\d{13}\s", ln):
            continue

        parts = ln.split()
        num_indices = [
            idx for idx, tok in enumerate(parts) if re.fullmatch(r"[\d.,]+", tok)
        ]
        if len(num_indices) < 6:
            continue

        qty_idx = num_indices[-6]
        p_bruto_idx = num_indices[-4]

        qty_str = parts[qty_idx]
        p_bruto_raw = parts[p_bruto_idx]
        precio = p_bruto_raw.replace(".", "").replace(",", ".")

        desc_tokens = parts[5:qty_idx]
        descripcion = " ".join(desc_tokens)

        extra_desc = ""
        if i + 1 < len(lines):
            next_ln = lines[i + 1]
            if (
                not re.match(r"^\d+\s+\d{13}\s", next_ln)
                and "WOMAN FIESTA" not in next_ln
                and not next_ln.startswith("Nº ")
            ):
                extra_desc = next_ln.strip()

        if extra_desc:
            descripcion = f"{descripcion} {extra_desc}"

        j = i + 1 + (1 if extra_desc else 0)
        modelo = ""
        color = ""
        talla_cod = ""
        talla = ""

        if j < len(lines):
            info_ln = lines[j]
            info_parts = info_ln.split()

            if len(info_parts) >= 2 and re.fullmatch(r"[A-Z0-9]+", info_parts[0]):
                modelo = info_parts[0]

                talla_idx = None
                talla_cod_found = None
                for idx in range(len(info_parts) - 1, 0, -1):
                    m = re.search(r"(\d{3})$", info_parts[idx])
                    if m:
                        talla_idx = idx
                        talla_cod_found = m.group(1)
                        break

                color_tokens = []
                if talla_idx is not None:
                    talla_cod = talla_cod_found
                    talla = SIZE_CODE_MAP.get(talla_cod, "")
                    base_color = re.sub(r"\d{3}$", "", info_parts[talla_idx])

                    color_tokens = info_parts[1:talla_idx]
                    if base_color:
                        color_tokens.append(base_color)
                else:
                    color_tokens = info_parts[1:]

                color = " ".join(color_tokens)

        rows.append(
            {
                "TIPO": tipo,
                "N_PEDIDO": n_pedido,
                "DEPARTAMENTO": departamento,
                "DESCRIPCION": descripcion,
                "MODELO": modelo,
                "COLOR": color,
                "PRECIO": precio,
                "FECHA_ENTREGA": fecha_entrega,
                "SUC_ENTREGA": suc_entrega,
                "TALLA_COD": talla_cod,
                "TALLA": talla,
                "UNIDADES": qty_str,
            }
        )

    return rows


def parse_pdf_eci_bytes(pdf_bytes: bytes) -> pd.DataFrame:
    all_rows = []
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            all_rows.extend(parse_page_eci(page.extract_text() or ""))

    if not all_rows:
        return pd.DataFrame()

    df = pd.DataFrame(all_rows)
    df["UNIDADES"] = (
        pd.to_numeric(df["UNIDADES"], errors="coerce").fillna(0).astype(int)
    )
    df["TALLA"] = df["TALLA"].fillna("").astype(str)

    group_cols_base = [
        "TIPO",
        "N_PEDIDO",
        "DEPARTAMENTO",
        "DESCRIPCION",
        "MODELO",
        "COLOR",
        "PRECIO",
        "FECHA_ENTREGA",
        "SUC_ENTREGA",
    ]

    df_grouped = df.groupby(group_cols_base + ["TALLA"], as_index=False)[
        "UNIDADES"
    ].sum()

    wide = df_grouped.pivot_table(
        index=group_cols_base, columns="TALLA", values="UNIDADES", fill_value=0
    ).reset_index()

    wide.columns.name = None

    for t in TALLAS:
        if t not in wide.columns:
            wide[t] = 0
        wide[t] = pd.to_numeric(wide[t], errors="coerce").fillna(0).astype(int)

    wide["TOTAL_UNIDADES"] = wide[TALLAS].sum(axis=1).astype(int)

    cols_order = group_cols_base + ["TOTAL_UNIDADES"] + TALLAS
    cols_order = [c for c in cols_order if c in wide.columns]
    return wide[cols_order]


# ===================== MAPA DE PAGINAS (PDF POR MODELO) =====================


def build_eurofiel_model_page_map(df: pd.DataFrame, pdf_bytes: bytes):
    if df.empty:
        return {}, 0

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        page_texts = [page.extract_text() or "" for page in pdf.pages]

    keys = set()
    for _, row in df.iterrows():
        modelo = str(row.get("MODELO") or "").strip()
        patron = str(row.get("PATRON") or "").strip()
        if modelo:
            keys.add((modelo, patron))

    model_map: Dict[Tuple[str, str], Set[int]] = {k: set() for k in keys}
    last_keys: Set[Tuple[str, str]] = set()

    for page_num, text in enumerate(page_texts, start=1):
        current: Set[Tuple[str, str]] = set()
        for modelo, patron in keys:
            if modelo in text and (not patron or patron in text):
                model_map[(modelo, patron)].add(page_num)
                current.add((modelo, patron))
        if current:
            last_keys = current
        else:
            for k in last_keys:
                model_map[k].add(page_num)

    return model_map, len(page_texts)


def build_eci_model_page_map(df: pd.DataFrame, pdf_bytes: bytes):
    if df.empty:
        return {}, 0

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        page_texts = [page.extract_text() or "" for page in pdf.pages]

    modelos = [str(m).strip() for m in df["MODELO"].dropna().unique() if str(m).strip()]
    model_map: Dict[str, Set[int]] = {m: set() for m in modelos}
    last_models: Set[str] = set()

    for page_num, text in enumerate(page_texts, start=1):
        current: Set[str] = set()
        for modelo in modelos:
            if modelo in text:
                model_map[modelo].add(page_num)
                current.add(modelo)
        if current:
            last_models = current
        else:
            for m in last_models:
                model_map[m].add(page_num)

    return model_map, len(page_texts)


# ===================== CREAR CARPETAS (LOCAL) =====================


def create_eurofiel_folders_and_pdfs_local(
    df: pd.DataFrame, pdf_bytes: bytes, base_dir: str
):
    os.makedirs(base_dir, exist_ok=True)
    model_map, total_pages = build_eurofiel_model_page_map(df, pdf_bytes)
    if not model_map:
        return

    reader = PdfReader(BytesIO(pdf_bytes))

    for (modelo, patron), pages in model_map.items():
        if not pages:
            continue

        folder_name = safe_fs_name(f"{modelo}_{patron}")
        folder_path = os.path.join(base_dir, folder_name)
        os.makedirs(folder_path, exist_ok=True)

        writer = PdfWriter()
        for p in sorted(pages):
            if 1 <= p <= total_pages:
                writer.add_page(reader.pages[p - 1])

        pdf_out_path = os.path.join(folder_path, f"{folder_name}.pdf")
        with open(pdf_out_path, "wb") as f_out:
            writer.write(f_out)

        sub_df = df[
            (df["MODELO"].astype(str).str.strip() == str(modelo).strip())
            & (df["PATRON"].astype(str).str.strip() == str(patron).strip())
        ].copy()

        if sub_df.empty:
            continue

        if "TOTAL_UNIDADES" in sub_df.columns:
            sub_df["TOTAL_UNIDADES"] = (
                pd.to_numeric(sub_df["TOTAL_UNIDADES"], errors="coerce")
                .fillna(0)
                .astype(int)
            )
            total_unidades = int(sub_df["TOTAL_UNIDADES"].sum())
        else:
            total_unidades = None

        for t in TALLAS:
            if t in sub_df.columns:
                sub_df[t] = (
                    pd.to_numeric(sub_df[t], errors="coerce").fillna(0).astype(int)
                )

        active_sizes_model = [
            t for t in TALLAS if t in sub_df.columns and int(sub_df[t].sum()) > 0
        ]
        zero_sizes_model = [
            t for t in TALLAS if t in sub_df.columns and t not in active_sizes_model
        ]
        if zero_sizes_model:
            sub_df = sub_df.drop(columns=zero_sizes_model)

        total_pedidos = (
            sub_df["PEDIDO"].nunique() if "PEDIDO" in sub_df.columns else None
        )

        cols = list(sub_df.columns)
        total_row = {col: "" for col in cols}
        if cols:
            total_row[cols[0]] = "TOTAL"
        if "PEDIDO" in cols and total_pedidos is not None:
            total_row["PEDIDO"] = total_pedidos
        if "TOTAL_UNIDADES" in cols and total_unidades is not None:
            total_row["TOTAL_UNIDADES"] = total_unidades
        for t in active_sizes_model:
            if t in cols:
                total_row[t] = int(sub_df[t].sum())

        sub_df_with_total = pd.concat(
            [sub_df, pd.DataFrame([total_row])], ignore_index=True
        )

        excel_out_path = os.path.join(folder_path, f"{folder_name}.xlsx")
        with pd.ExcelWriter(excel_out_path, engine="openpyxl") as writer_x:
            sub_df_with_total.to_excel(writer_x, index=False, sheet_name="Resumen")
            style_workbook_with_borders_and_headers(writer_x.book)

            # Pintar filas por modelo si existe la columna y la hoja
            for sheet_name in ("Pedidos", "Resumen"):
                if sheet_name in writer_x.book.sheetnames:
                    ws = writer_x.book[sheet_name]
                    apply_model_row_colors(ws, "MODELO")


def create_eci_folders_and_pdfs_local(
    df: pd.DataFrame, pdf_bytes: bytes, base_dir: str
):
    os.makedirs(base_dir, exist_ok=True)
    model_map, total_pages = build_eci_model_page_map(df, pdf_bytes)
    if not model_map:
        return

    reader = PdfReader(BytesIO(pdf_bytes))

    for modelo, pages in model_map.items():
        if not pages:
            continue

        folder_name = safe_fs_name(modelo)
        folder_path = os.path.join(base_dir, folder_name)
        os.makedirs(folder_path, exist_ok=True)

        writer = PdfWriter()
        for p in sorted(pages):
            if 1 <= p <= total_pages:
                writer.add_page(reader.pages[p - 1])

        pdf_out_path = os.path.join(folder_path, f"{folder_name}.pdf")
        with open(pdf_out_path, "wb") as f_out:
            writer.write(f_out)

        sub_df = df[df["MODELO"].astype(str).str.strip() == str(modelo).strip()].copy()
        if sub_df.empty:
            continue

        if "TOTAL_UNIDADES" in sub_df.columns:
            sub_df["TOTAL_UNIDADES"] = (
                pd.to_numeric(sub_df["TOTAL_UNIDADES"], errors="coerce")
                .fillna(0)
                .astype(int)
            )
            total_unidades = int(sub_df["TOTAL_UNIDADES"].sum())
        else:
            total_unidades = None

        for t in TALLAS:
            if t in sub_df.columns:
                sub_df[t] = (
                    pd.to_numeric(sub_df[t], errors="coerce").fillna(0).astype(int)
                )

        active_sizes_model = [
            t for t in TALLAS if t in sub_df.columns and int(sub_df[t].sum()) > 0
        ]
        zero_sizes_model = [
            t for t in TALLAS if t in sub_df.columns and t not in active_sizes_model
        ]
        if zero_sizes_model:
            sub_df = sub_df.drop(columns=zero_sizes_model)

        total_pedidos = (
            sub_df["N_PEDIDO"].nunique() if "N_PEDIDO" in sub_df.columns else None
        )

        cols = list(sub_df.columns)
        total_row = {col: "" for col in cols}
        if cols:
            total_row[cols[0]] = "TOTAL"
        if "N_PEDIDO" in cols and total_pedidos is not None:
            total_row["N_PEDIDO"] = total_pedidos
        if "TOTAL_UNIDADES" in cols and total_unidades is not None:
            total_row["TOTAL_UNIDADES"] = total_unidades
        for t in active_sizes_model:
            if t in cols:
                total_row[t] = int(sub_df[t].sum())

        sub_df_with_total = pd.concat(
            [sub_df, pd.DataFrame([total_row])], ignore_index=True
        )

        excel_out_path = os.path.join(folder_path, f"{folder_name}.xlsx")
        with pd.ExcelWriter(excel_out_path, engine="openpyxl") as writer_x:
            sub_df_with_total.to_excel(writer_x, index=False, sheet_name="Resumen")
            style_workbook_with_borders_and_headers(writer_x.book)

            for sheet_name in ("Pedidos", "Resumen"):
                if sheet_name in writer_x.book.sheetnames:
                    ws = writer_x.book[sheet_name]
                    apply_model_row_colors(ws, "MODELO")


# ===================== TXT: EUROFIEL =====================


def split_ediwin_txt_files_per_model_eurofiel(
    base_dir: str,
    df: pd.DataFrame,
    linped_bytes: bytes,
    linped_name: str,
    cabped_bytes: bytes | None = None,
    cabped_name: str | None = None,
    locped_bytes: bytes | None = None,
    locped_name: str | None = None,
    obsped_bytes: bytes | None = None,
    obsped_name: str | None = None,
    obslped_bytes: bytes | None = None,
    obslped_name: str | None = None,
    encoding: str = "latin-1",
    max_model_length: int | None = None,
):
    if linped_bytes is None:
        raise ValueError("LINPED es obligatorio")

    def decode_lines(b: bytes | None) -> list[str]:
        if b is None:
            return []
        text = b.decode(encoding, errors="ignore")
        return text.splitlines(keepends=True)

    lin_lines = decode_lines(linped_bytes)
    cab_lines = decode_lines(cabped_bytes)
    loc_lines = decode_lines(locped_bytes)
    obs_lines = decode_lines(obsped_bytes)
    obsl_lines = decode_lines(obslped_bytes)

    model_changes: dict[str, str] = {}

    def shorten_cod_prov_for_sage(cod_prov: str, max_len: int) -> str:
        cod_prov = cod_prov.strip()
        if max_len is None or max_len <= 0:
            return cod_prov

        if "/" in cod_prov:
            base, talla = cod_prov.rsplit("/", 1)
            base = base.strip()
            original_base = base

            new_base = base[:max_len] if len(base) > max_len else base
            if new_base != original_base:
                model_changes.setdefault(original_base, new_base)

            return f"{new_base}/{talla}"

        original_base = cod_prov
        if len(cod_prov) > max_len:
            new_base = cod_prov[:max_len]
            if new_base != original_base:
                model_changes.setdefault(original_base, new_base)
            return new_base

        return cod_prov

    def adjust_lin_line_model(line: str) -> str:
        if max_model_length is None:
            return line
        if not line.strip():
            return line

        newline = ""
        if line.endswith("\r\n"):
            newline = "\r\n"
        elif line.endswith("\n"):
            newline = "\n"

        core = line.rstrip("\r\n")
        parts = core.split(";")
        if len(parts) < 7:
            return line

        cod_prov = parts[6].strip()
        if not cod_prov:
            return line

        parts[6] = shorten_cod_prov_for_sage(cod_prov, max_model_length)
        return ";".join(parts) + newline

    pedido_to_folders: dict[str, set[str]] = defaultdict(set)

    for line in lin_lines:
        if not line.strip():
            continue
        parts = line.split(";")
        if len(parts) < 7:
            continue

        pedido_int = parts[0].strip()
        cod_cli = parts[5].strip()
        cod_prov = parts[6].strip()

        if not pedido_int or not cod_cli or not cod_prov:
            continue

        patron = re.sub(r"/[^/]+$", "", cod_cli)
        modelo = re.sub(r"/[^/]+$", "", cod_prov)
        folder_name = safe_fs_name(f"{modelo}_{patron}")
        pedido_to_folders[pedido_int].add(folder_name)

    if not pedido_to_folders:
        return model_changes

    def distribute_by_folder(lines: list[str]) -> dict[str, list[str]]:
        out: dict[str, list[str]] = defaultdict(list)
        for line in lines:
            if not line.strip():
                continue
            parts = line.split(";")
            if not parts:
                continue
            pedido_int = parts[0].strip()
            if not pedido_int:
                continue
            folders = pedido_to_folders.get(pedido_int)
            if not folders:
                continue
            for folder in folders:
                out[folder].append(line)
        return out

    lin_by_folder = distribute_by_folder(lin_lines)
    cab_by_folder = distribute_by_folder(cab_lines)
    loc_by_folder = distribute_by_folder(loc_lines)
    obs_by_folder = distribute_by_folder(obs_lines)
    obsl_by_folder = distribute_by_folder(obsl_lines)

    all_folders = sorted(
        {folder for folders in pedido_to_folders.values() for folder in folders}
    )

    os.makedirs(base_dir, exist_ok=True)

    for folder in all_folders:
        folder_path = os.path.join(base_dir, folder)
        os.makedirs(folder_path, exist_ok=True)

        if linped_name:
            lin_path = os.path.join(folder_path, linped_name)
            with open(lin_path, "w", encoding=encoding, newline="") as f:
                raw_lines = lin_by_folder.get(folder, [])
                adj_lines = [adjust_lin_line_model(line) for line in raw_lines]
                f.writelines(adj_lines)

        if cabped_name:
            with open(
                os.path.join(folder_path, cabped_name),
                "w",
                encoding=encoding,
                newline="",
            ) as f:
                f.writelines(cab_by_folder.get(folder, []))

        if locped_name:
            with open(
                os.path.join(folder_path, locped_name),
                "w",
                encoding=encoding,
                newline="",
            ) as f:
                f.writelines(loc_by_folder.get(folder, []))

        if obsped_name:
            with open(
                os.path.join(folder_path, obsped_name),
                "w",
                encoding=encoding,
                newline="",
            ) as f:
                f.writelines(obs_by_folder.get(folder, []))

        if obslped_name:
            with open(
                os.path.join(folder_path, obslped_name),
                "w",
                encoding=encoding,
                newline="",
            ) as f:
                f.writelines(obsl_by_folder.get(folder, []))

    return model_changes


# ===================== TXT: ECI =====================


def split_ediwin_txt_files_per_model(
    base_dir: str,
    linped_bytes: bytes,
    linped_name: str,
    cabped_bytes: bytes | None = None,
    cabped_name: str | None = None,
    locped_bytes: bytes | None = None,
    locped_name: str | None = None,
    obsped_bytes: bytes | None = None,
    obsped_name: str | None = None,
    obslped_bytes: bytes | None = None,
    obslped_name: str | None = None,
    encoding: str = "latin-1",
):
    if linped_bytes is None:
        raise ValueError("LINPED es obligatorio")

    def decode_lines(b: bytes | None) -> list[str]:
        if b is None:
            return []
        text = b.decode(encoding, errors="ignore")
        return text.splitlines(keepends=True)

    lin_lines = decode_lines(linped_bytes)
    cab_lines = decode_lines(cabped_bytes)
    loc_lines = decode_lines(locped_bytes)
    obs_lines = decode_lines(obsped_bytes)
    obsl_lines = decode_lines(obslped_bytes)

    pedido_to_folders: dict[str, set[str]] = defaultdict(set)

    for line in lin_lines:
        if not line.strip():
            continue

        parts = line.split(";")
        if len(parts) < 7:
            continue

        pedido_int = parts[0].strip()
        cod_prov = parts[6].strip()
        if not pedido_int or not cod_prov:
            continue

        modelo = re.sub(r"/[^/]+$", "", cod_prov).strip()
        folder_name = safe_fs_name(modelo or "SIN_MODELO")
        pedido_to_folders[pedido_int].add(folder_name)

    if not pedido_to_folders:
        return

    def distribute_by_folder(lines: list[str]) -> dict[str, list[str]]:
        out: dict[str, list[str]] = defaultdict(list)
        for line in lines:
            if not line.strip():
                continue
            parts = line.split(";")
            if not parts:
                continue
            pedido_int = parts[0].strip()
            if not pedido_int:
                continue

            folders = pedido_to_folders.get(pedido_int)
            if not folders:
                continue

            for folder in folders:
                out[folder].append(line)

        return out

    lin_by_folder = distribute_by_folder(lin_lines)
    cab_by_folder = distribute_by_folder(cab_lines)
    loc_by_folder = distribute_by_folder(loc_lines)
    obs_by_folder = distribute_by_folder(obs_lines)
    obsl_by_folder = distribute_by_folder(obsl_lines)

    all_folders = sorted(
        {folder for folders in pedido_to_folders.values() for folder in folders}
    )

    os.makedirs(base_dir, exist_ok=True)

    for folder in all_folders:
        folder_path = os.path.join(base_dir, folder)
        os.makedirs(folder_path, exist_ok=True)

        if linped_name:
            with open(
                os.path.join(folder_path, linped_name),
                "w",
                encoding=encoding,
                newline="",
            ) as f:
                f.writelines(lin_by_folder.get(folder, []))

        if cabped_name:
            with open(
                os.path.join(folder_path, cabped_name),
                "w",
                encoding=encoding,
                newline="",
            ) as f:
                f.writelines(cab_by_folder.get(folder, []))

        if locped_name:
            with open(
                os.path.join(folder_path, locped_name),
                "w",
                encoding=encoding,
                newline="",
            ) as f:
                f.writelines(loc_by_folder.get(folder, []))

        if obsped_name:
            with open(
                os.path.join(folder_path, obsped_name),
                "w",
                encoding=encoding,
                newline="",
            ) as f:
                f.writelines(obs_by_folder.get(folder, []))

        if obslped_name:
            with open(
                os.path.join(folder_path, obslped_name),
                "w",
                encoding=encoding,
                newline="",
            ) as f:
                f.writelines(obsl_by_folder.get(folder, []))


# ===================== EXPORT FULL (ZIP) =====================


def export_outputs_zip_full(
    tipo: str,
    pdf_bytes: bytes,
    outzip_path: str,
    include_model_folders: bool,
    # TXT (opcionales)
    linped_bytes: bytes | None = None,
    linped_name: str | None = None,
    cabped_bytes: bytes | None = None,
    cabped_name: str | None = None,
    locped_bytes: bytes | None = None,
    locped_name: str | None = None,
    obsped_bytes: bytes | None = None,
    obsped_name: str | None = None,
    obslped_bytes: bytes | None = None,
    obslped_name: str | None = None,
    recortar_modelo_sage: bool = True,  # Eurofiel only
    sage_max_len: int = 20,
):
    """
    Replica funcionalidad Streamlit:
      - parse PDF -> df
      - export Excel + CSV
      - include_model_folders: crea carpetas por modelo con PDF+XLSX
      - TXT: si vienen, los reparte por modelo (Eurofiel o ECI)
      - empaqueta todo en ZIP (outzip_path)
    """
    tipo = tipo.upper()
    if tipo not in ("EUROFIEL", "ECI"):
        raise ValueError("tipo debe ser EUROFIEL o ECI")

    outzip = Path(outzip_path)
    work_dir = outzip.parent / f"_work_{outzip.stem}"
    if work_dir.exists():
        # limpieza best-effort
        for p in work_dir.rglob("*"):
            if p.is_file():
                try:
                    p.unlink()
                except Exception:
                    pass
    work_dir.mkdir(parents=True, exist_ok=True)

    # 1) Parse PDF
    if tipo == "EUROFIEL":
        df = parse_pdf_eurofiel_bytes(pdf_bytes)
    else:
        df = parse_pdf_eci_bytes(pdf_bytes)

    if df.empty:
        raise ValueError("No se han detectado pedidos en el PDF.")

    # TOTAL_UNIDADES numérico si existe
    if "TOTAL_UNIDADES" in df.columns:
        df["TOTAL_UNIDADES"] = (
            pd.to_numeric(df["TOTAL_UNIDADES"], errors="coerce").fillna(0).astype(int)
        )

    # 2) Quitar tallas a 0 para export visible
    active_sizes = get_active_sizes(df)
    zero_size_cols = [t for t in TALLAS if t in df.columns and t not in active_sizes]
    df_display = df.drop(columns=zero_size_cols) if zero_size_cols else df.copy()

    # 3) CSV
    csv_path = work_dir / f"{tipo.lower()}_resumen_pedidos.csv"
    csv_path.write_bytes(df_display.to_csv(index=False).encode("utf-8"))

    # 4) Excel + resúmenes
    excel_path = work_dir / f"{tipo.lower()}_resumen_pedidos.xlsx"

    if tipo == "EUROFIEL":
        resumen = (
            df.groupby("MODELO", dropna=False)
            .agg(
                PEDIDOS=("PEDIDO", "nunique"),
                UNIDADES_TOTALES=("TOTAL_UNIDADES", "sum"),
            )
            .reset_index()
            .sort_values("PEDIDOS", ascending=False)
        )
        resumen_xlsx = resumen.copy()
        total_row = {
            "MODELO": "TOTAL",
            "PEDIDOS": int(resumen_xlsx["PEDIDOS"].sum()),
            "UNIDADES_TOTALES": int(resumen_xlsx["UNIDADES_TOTALES"].sum()),
        }
        resumen_xlsx = pd.concat(
            [resumen_xlsx, pd.DataFrame([total_row])], ignore_index=True
        )

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer_x:
            df_display.to_excel(writer_x, sheet_name="Pedidos", index=False)
            ws_pedidos = writer_x.book["Pedidos"]
            apply_model_row_colors(ws_pedidos, "MODELO")

            resumen_xlsx.to_excel(
                writer_x, sheet_name="Resumen por modelo", index=False
            )
            style_workbook_with_borders_and_headers(writer_x.book)

        if include_model_folders:
            create_eurofiel_folders_and_pdfs_local(
                df, pdf_bytes, str(work_dir / "por_modelo")
            )

    else:
        resumen_mc = (
            df.groupby(["MODELO", "COLOR"], dropna=False)
            .agg(
                PEDIDOS=("N_PEDIDO", "nunique"),
                UNIDADES_TOTALES=("TOTAL_UNIDADES", "sum"),
            )
            .reset_index()
            .sort_values("PEDIDOS", ascending=False)
        )
        resumen_m = (
            df.groupby("MODELO", dropna=False)
            .agg(
                PEDIDOS=("N_PEDIDO", "nunique"),
                UNIDADES_TOTALES=("TOTAL_UNIDADES", "sum"),
            )
            .reset_index()
            .sort_values("PEDIDOS", ascending=False)
        )

        resumen_mc_xlsx = resumen_mc.copy()
        total_mc = {
            "MODELO": "TOTAL",
            "COLOR": "",
            "PEDIDOS": int(resumen_mc_xlsx["PEDIDOS"].sum()),
            "UNIDADES_TOTALES": int(resumen_mc_xlsx["UNIDADES_TOTALES"].sum()),
        }
        resumen_mc_xlsx = pd.concat(
            [resumen_mc_xlsx, pd.DataFrame([total_mc])], ignore_index=True
        )

        resumen_m_xlsx = resumen_m.copy()
        total_m = {
            "MODELO": "TOTAL",
            "PEDIDOS": int(resumen_m_xlsx["PEDIDOS"].sum()),
            "UNIDADES_TOTALES": int(resumen_m_xlsx["UNIDADES_TOTALES"].sum()),
        }
        resumen_m_xlsx = pd.concat(
            [resumen_m_xlsx, pd.DataFrame([total_m])], ignore_index=True
        )

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer_x:
            df_display.to_excel(writer_x, sheet_name="Pedidos", index=False)
            resumen_mc_xlsx.to_excel(
                writer_x, sheet_name="Resumen modelo+color", index=False
            )
            resumen_m_xlsx.to_excel(writer_x, sheet_name="Resumen modelo", index=False)
            style_workbook_with_borders_and_headers(writer_x.book)

        if include_model_folders:
            create_eci_folders_and_pdfs_local(
                df, pdf_bytes, str(work_dir / "por_modelo")
            )

    # 5) TXT: repartir por modelo, dentro de work_dir/txt_por_modelo
    model_changes = None

    if linped_bytes and linped_name:
        txt_dir = work_dir / "txt_por_modelo"
        if tipo == "EUROFIEL":
            max_len = sage_max_len if recortar_modelo_sage else None
            model_changes = split_ediwin_txt_files_per_model_eurofiel(
                base_dir=str(txt_dir),
                df=df,
                linped_bytes=linped_bytes,
                linped_name=linped_name,
                cabped_bytes=cabped_bytes,
                cabped_name=cabped_name,
                locped_bytes=locped_bytes,
                locped_name=locped_name,
                obsped_bytes=obsped_bytes,
                obsped_name=obsped_name,
                obslped_bytes=obslped_bytes,
                obslped_name=obslped_name,
                max_model_length=max_len,
            )
        else:
            split_ediwin_txt_files_per_model(
                base_dir=str(txt_dir),
                linped_bytes=linped_bytes,
                linped_name=linped_name,
                cabped_bytes=cabped_bytes,
                cabped_name=cabped_name,
                locped_bytes=locped_bytes,
                locped_name=locped_name,
                obsped_bytes=obsped_bytes,
                obsped_name=obsped_name,
                obslped_bytes=obslped_bytes,
                obslped_name=obslped_name,
            )

    # 6) Si hubo recortes Sage, guardar un CSV de cambios
    if model_changes:
        changes_path = work_dir / "sage_model_changes.csv"
        df_changes = pd.DataFrame(
            [
                {"MODELO_ORIGINAL": k, "MODELO_SAGE": v}
                for k, v in sorted(model_changes.items())
            ]
        )
        changes_path.write_bytes(df_changes.to_csv(index=False).encode("utf-8"))

    # 7) Zip final
    zip_dir(str(work_dir), outzip_path)

    return {
        "ok": True,
        "tipo": tipo,
        "zip": str(outzip_path),
        "has_txt": bool(linped_bytes and linped_name),
        "has_model_folders": include_model_folders,
        "sage_changes_count": len(model_changes) if model_changes else 0,
    }


# ------------------------------------------------------------
# Compatibilidad hacia atrás:
# algunos callers (cli.py / integraciones) esperaban export_outputs_zip
# y ahora el nombre real es export_outputs_zip_full.
# Mantener alias evita roturas y nos permite refactorizar sin dolor.
# ------------------------------------------------------------
export_outputs_zip = export_outputs_zip_full
