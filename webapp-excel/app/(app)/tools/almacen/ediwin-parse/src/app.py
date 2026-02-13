import os
import re
from io import BytesIO

import streamlit as st
import pdfplumber
import pandas as pd
from openpyxl.styles import Border, Side, PatternFill, Font
from pypdf import PdfReader, PdfWriter
from collections import defaultdict


# Rutas de salida locales demo
EUROFIEL_BASE_DIR = r"./data/ediwin/out/eurofiel"
ECI_BASE_DIR      = r"./data/ediwin/out/eci"


# ============= PARSER EUROFIEL =============

def split_orders(full_text: str):
    """
    Divide el texto completo del PDF en bloques,
    cada uno correspondiente a un pedido (PEDIDO / REEMPLAZO / ANULACIÓN).
    """
    matches = list(re.finditer(r"Nº Pedido\s*:", full_text))
    chunks = []

    for i, m in enumerate(matches):
        start = m.start()
        # Buscamos la línea anterior para incluir el TIPO (PEDIDO, ANULACIÓN PEDIDO…)
        prev_nl = full_text.rfind("\n", 0, start)
        if prev_nl != -1:
            prev_prev_nl = full_text.rfind("\n", 0, prev_nl)
            order_start = prev_prev_nl + 1 if prev_prev_nl != -1 else 0
        else:
            order_start = 0

        end = matches[i + 1].start() if i + 1 < len(matches) else len(full_text)
        chunk = full_text[order_start:end].strip()
        chunks.append(chunk)

    return chunks


def parse_detail_line_eurofiel(line: str):
    """
    Parsea una línea de detalle de artículo Eurofiel.

    Ej:
      1 8447571299747 3RC240/NARANJA/XS 0863769/66/01 1 50 50 0 EUR

    Devuelve (MODELO, PATRON, TALLA, UNIDADES, PRECIO) o None si la línea no es de detalle.
    """
    parts = line.split()
    if len(parts) < 8:
        return None

    # nº línea
    if not parts[0].isdigit():
        return None

    # EAN13
    if not re.fullmatch(r"\d{13}", parts[1]):
        return None

    # Buscamos el código cliente con formato 0000000/00/00
    cli_idx = None
    for i in range(2, len(parts)):
        if re.fullmatch(r"\d+/\d+/\d+", parts[i]):
            cli_idx = i
            break

    if cli_idx is None or cli_idx + 4 >= len(parts):
        return None

    cod_prov_full = " ".join(parts[2:cli_idx])   # 3RC240/NARANJA/XS
    cod_cli_full = parts[cli_idx]               # 0863769/66/01

    # campos numéricos después del código cliente:
    # ... cod_cli  QTY  P_BRUTO  P_NETO  DTO  EUR
    qty = 0
    try:
        qty = int(parts[cli_idx + 1])
    except (ValueError, IndexError):
        qty = 0

    p_neto = parts[cli_idx + 3]  # como tenías antes
    precio = p_neto.replace(",", ".")

    # MODELO = Cod Proveedor/Color (quitamos talla)
    modelo = re.sub(r"/[^/]+$", "", cod_prov_full)
    # PATRON = Cod Cliente/Color (quitamos talla)
    patron = re.sub(r"/[^/]+$", "", cod_cli_full)

    # Talla = última parte del código proveedor
    talla = cod_prov_full.split("/")[-1].strip().upper()

    return modelo, patron, talla, qty, precio

from collections import defaultdict

def parse_order_eurofiel(order_text: str):
    """
    Parsea un bloque de texto correspondiente a un solo pedido Eurofiel.
    Ahora además acumula unidades por talla (XXS, XS, S, M, L, XL, XXL).
    """
    lines = [ln for ln in order_text.splitlines() if ln.strip()]
    first_line = lines[0].strip() if lines else ""
    tipo = first_line  # PEDIDO / REEMPLAZO PEDIDO / ANULACIÓN PEDIDO

    def search(pattern: str):
        m = re.search(pattern, order_text)
        return m.group(1).strip() if m else ""

    pedido = search(r"Nº Pedido\s*:\s*(\S+)")
    fecha_entrega = search(r"Fecha Entrega\s*:\s*(\d{2}/\d{2}/\d{4})")

    pais = ""
    m_pais = re.search(r"País:\s*\([^)]*\)\s*([A-ZÁÉÍÓÚÜÑ ]+)", order_text)
    if m_pais:
        pais = m_pais.group(1).strip()

    descripcion = search(r"Descripción:\s*(.+)")
    total_unidades = search(r"Total Unidades\s+(\d+)")

    modelo = ""
    patron = ""
    precio = ""

    # contador de tallas
    tallas_count = defaultdict(int)
    primer_detalle = True

    for ln in lines:
        parsed = parse_detail_line_eurofiel(ln)
        if parsed:
            m, p, talla, qty, prec = parsed

            # fijamos modelo/patrón/precio a partir de la primera línea de detalle
            if primer_detalle:
                modelo, patron, precio = m, p, prec
                primer_detalle = False

            # acumulamos unidades por talla
            talla_norm = talla.upper()
            tallas_count[talla_norm] += qty

    row = {
        "TIPO": tipo,
        "PEDIDO": pedido,
        "FECHA_ENTREGA": fecha_entrega,
        "PAIS": pais,
        "DESCRIPCION": descripcion,
        "MODELO": modelo,
        "PATRON": patron,
        "PRECIO": precio,
        "TOTAL_UNIDADES": total_unidades,
    }

    # Añadimos columnas de tallas
    for t in TALLAS:
        row[t] = tallas_count.get(t, 0)

    return row



def parse_pdf_eurofiel_bytes(pdf_bytes: bytes):
    """
    Parsea un PDF Eurofiel que viene en memoria (subido por web).
    """
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    orders = split_orders(full_text)
    rows = [parse_order_eurofiel(o) for o in orders]
    rows = [r for r in rows if r.get("PEDIDO")]  # limpia ruido
    return pd.DataFrame(rows)

# ============= PARSER ECI =============

def parse_page_eci(text: str):
    """
    Parsea una página de pedido de ECI.
    Devuelve una lista de dicts, una fila por (pedido, modelo, color, talla).
    """

    # Limpiamos líneas
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    def search(pattern: str):
        m = re.search(pattern, text)
        return m.group(1).strip() if m else ""

    # TIPO: Pedido / Reposición / Anulación Pedido...
    tipo = ""
    for ln in lines:
        low = ln.lower()
        if low in ("pedido", "reposicion", "reposición", "anulacion pedido", "anulación pedido"):
            tipo = ln.upper()
            break

    # Cabecera
    n_pedido = search(r"Nº Pedido\s+(\d+)")
    departamento = search(r"Dpto\. venta\s+(\d+)")
    fecha_entrega = search(r"Fecha Entrega\s+(\d{2}/\d{2}/\d{4})")

    # Sucursal entrega (01 0050, 02 0062, etc.)
    suc_entrega = search(r"Sucursal Destino que Pide\s+([0-9 ]+)\s+[A-ZÁÉÍÓÚÜÑ]")
    if not suc_entrega:
        suc_entrega = search(r"Sucursal de Entrega\s+([0-9 ]+)\s+[A-ZÁÉÍÓÚÜÑ]")

    rows = []

    for i, ln in enumerate(lines):
        # Línea de detalle principal (nº + EAN13 + resto)
        if not re.match(r"^\d+\s+\d{13}\s", ln):
            continue

        parts = ln.split()

        # Índices de tokens numéricos (nos sirven para localizar QTY y precios)
        num_indices = [idx for idx, tok in enumerate(parts)
                       if re.fullmatch(r"[\d.,]+", tok)]
        if len(num_indices) < 6:
            continue

        # Patrón: ... DESCRIPCION QTY 1 P_BRUTO P_NETO PVP NETO_LINEA
        qty_idx = num_indices[-6]
        p_bruto_idx = num_indices[-4]

        qty_str = parts[qty_idx]
        p_bruto_raw = parts[p_bruto_idx]

        # Normalizamos precio: 53,000 → 53.000
        precio = p_bruto_raw.replace(".", "").replace(",", ".")

        # Descripción: después de los 3 códigos (serie + ref + colorcode)
        desc_tokens = parts[5:qty_idx]
        descripcion = " ".join(desc_tokens)

        # Línea siguiente puede ser segunda parte de la descripción
        extra_desc = ""
        if i + 1 < len(lines):
            next_ln = lines[i + 1]
            if (
                not re.match(r"^\d+\s+\d{13}\s", next_ln)  # no es otra línea de detalle
                and "WOMAN FIESTA" not in next_ln          # no es la línea de marca
                and not next_ln.startswith("Nº ")          # no es cabecera
            ):
                extra_desc = next_ln.strip()

        if extra_desc:
            descripcion = f"{descripcion} {extra_desc}"

        # Línea de modelo/color/talla:
        # si hay extra_desc, está 2 líneas debajo; si no, 1 línea debajo
        j = i + 1 + (1 if extra_desc else 0)
        modelo = ""
        color = ""
        talla_cod = ""
        talla = ""

        if j < len(lines):
            info_ln = lines[j]
            info_parts = info_ln.split()

            # Ej: "67D2677 980 P NEGR FLOR003 3"
            if len(info_parts) >= 2 and re.fullmatch(r"[A-Z0-9]+", info_parts[0]):
                modelo = info_parts[0]

                # Buscamos la talla como 3 dígitos al final de algún token (FLOR003 → 003)
                talla_idx = None
                talla_cod_found = None
                for idx in range(len(info_parts) - 1, 0, -1):
                    m = re.search(r"(\d{3})$", info_parts[idx])
                    if m:
                        talla_idx = idx
                        talla_cod_found = m.group(1)
                        break

                color_tokens = []
                if talla_idx is not None:
                    talla_cod = talla_cod_found
                    talla = SIZE_CODE_MAP.get(talla_cod, "")

                    # El token de color puede llevar pegada la talla: FLOR003 → FLOR
                    base_color = re.sub(r"\d{3}$", "", info_parts[talla_idx])

                    # Color = todo lo que hay entre modelo y el token con talla,
                    # más la parte de texto de ese token si queda algo
                    color_tokens = info_parts[1:talla_idx]
                    if base_color:
                        color_tokens.append(base_color)
                else:
                    # Si no encontramos talla, usamos todo lo demás como color
                    color_tokens = info_parts[1:]

                color = " ".join(color_tokens)

        rows.append({
            "TIPO": tipo,
            "N_PEDIDO": n_pedido,
            "DEPARTAMENTO": departamento,
            "DESCRIPCION": descripcion,
            "MODELO": modelo,
            "COLOR": color,
            "PRECIO": precio,
            "FECHA_ENTREGA": fecha_entrega,
            "SUC_ENTREGA": suc_entrega,
            "TALLA_COD": talla_cod,
            "TALLA": talla,
            "UNIDADES": qty_str,
        })

    return rows


def parse_pdf_eci_bytes(pdf_bytes: bytes) -> pd.DataFrame:
    """
    Abre un PDF de ECI (bytes) y devuelve un DataFrame con:
    - 1 fila por (PEDIDO + MODELO + COLOR + sucursal)
    - Columnas de tallas (XXS, XS, S, M, L, XL, ... 34, 36, 38, ...)
    - TOTAL_UNIDADES = suma de todas las tallas
    """
    all_rows = []

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            all_rows.extend(parse_page_eci(text))

    if not all_rows:
        return pd.DataFrame()

    df = pd.DataFrame(all_rows)

    # Unidades numéricas
    df["UNIDADES"] = pd.to_numeric(df["UNIDADES"], errors="coerce").fillna(0).astype(int)

    # Aseguramos que la talla tenga nombre (por si viene solo el código)
    df["TALLA"] = df["TALLA"].fillna("").astype(str)

    group_cols_base = [
        "TIPO",
        "N_PEDIDO",
        "DEPARTAMENTO",
        "DESCRIPCION",
        "MODELO",
        "COLOR",
        "PRECIO",
        "FECHA_ENTREGA",
        "SUC_ENTREGA",
    ]

    # Sumamos unidades por combinación + talla
    df_grouped = (
        df.groupby(group_cols_base + ["TALLA"], as_index=False)["UNIDADES"]
          .sum()
    )

    # Pivotamos tallas a columnas
    wide = df_grouped.pivot_table(
        index=group_cols_base,
        columns="TALLA",
        values="UNIDADES",
        fill_value=0,
    ).reset_index()

    wide.columns.name = None  # quitamos el nombre del índice de columnas

    # Nos aseguramos de que existan todas las columnas de TALLAS (aunque sea a 0)
    for t in TALLAS:
        if t not in wide.columns:
            wide[t] = 0

    # ✅ Cast a int todas las columnas de tallas (pivot_table las deja en float)
    for t in TALLAS:
        if t in wide.columns:
            wide[t] = pd.to_numeric(wide[t], errors="coerce").fillna(0).astype(int)

    # TOTAL_UNIDADES = suma de todas las tallas (como int)
    wide["TOTAL_UNIDADES"] = wide[TALLAS].sum(axis=1).astype(int)

    # Orden de columnas: cabecera + TOTAL_UNIDADES + tallas
    cols_order = group_cols_base + ["TOTAL_UNIDADES"] + TALLAS
    cols_order = [c for c in cols_order if c in wide.columns]
    wide = wide[cols_order]

    return wide



# ============= UTILIDADES COMUNES (COLORES + BORDES + TOTALES) =============

PALETTE = [
    "#fde2e4",
    "#bee1e6",
    "#e2f0cb",
    "#ffdfba",
    "#d0bdf4",
    "#c7f9cc",
    "#ffccd5",
    "#f1f0ff",
    "#e5f4e3",
    "#ffe5b4",
    "#e0bbff",
    "#caffbf",
    "#ffd6a5",
    "#bde0fe",
    "#ffafcc",
]

TALLAS = ["XXS", "XS", "S", "M", "L", "XL", "XXL", "34", "36", "38", "40", "42", "44", "46", "48"]

SIZE_CODE_MAP = {
    "001": "XXS",
    "002": "XS",
    "003": "S",
    "004": "M",
    "005": "L",
    "006": "XL",
    "007": "XXL",
    "034": "34",
    "036": "36",
    "038": "38",
    "040": "40",
    "042": "42",
    "044": "44",
    "046": "46",
    "048": "48",
}


def get_active_sizes(df: pd.DataFrame):
    """
    Devuelve la lista de tallas cuya suma total en el DataFrame > 0.
    Sirve para ocultar columnas de tallas que están todo a 0.
    """
    activas = []
    for t in TALLAS:
        if t in df.columns:
            total = pd.to_numeric(df[t], errors="coerce").fillna(0).sum()
            if total > 0:
                activas.append(t)
    return activas


def style_by_model(df: pd.DataFrame):
    modelos = df["MODELO"].fillna("").astype(str).unique()
    model_colors = {}
    for i, m in enumerate(modelos):
        if not m:
            continue
        color = PALETTE[i % len(PALETTE)]
        model_colors[m] = color

    def color_rows(row):
        color = model_colors.get(str(row["MODELO"]), "")
        if not color:
            return [""] * len(row)
        return [f"background-color: {color}"] * len(row)

    return df.style.apply(color_rows, axis=1)


def style_workbook_with_borders_and_headers(workbook):
    """
    Aplica:
    - Bordes finos a todas las celdas
    - Cabeceras en amarillo chillón + negrita (fila 1)
    - Filas cuyo primer valor sea 'TOTAL' → amarillo + negrita
    """
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="FFFF00")  # amarillo fuerte
    header_font = Font(bold=True)

    total_fill = PatternFill("solid", fgColor="FFFF00")
    total_font = Font(bold=True)

    for ws in workbook.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column

        # Bordes + cabecera
        for row in ws.iter_rows(min_row=1, max_row=max_row,
                                min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font

        # Totales (fila con primera celda == 'TOTAL')
        for row_idx in range(2, max_row + 1):
            first_cell = ws.cell(row=row_idx, column=1)
            if str(first_cell.value).strip().upper() == "TOTAL":
                for col_idx in range(1, max_col + 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.fill = total_fill
                    c.font = total_font

# ============= PARTIR PDF Y CREAR CARPETAS =============

def safe_fs_name(name: str) -> str:
    """
    Limpia un nombre para usarlo como nombre de carpeta/archivo en Windows.
    - Sustituye / y \ por -
    - Quita caracteres raros
    """
    name = name.replace("/", "-").replace("\\", "-")
    # Cualquier cosa que no sea letra, número, guion o guion bajo → _
    name = re.sub(r"[^\w\-]+", "_", name)
    return name.strip("_") or "SIN_NOMBRE"


def build_eurofiel_model_page_map(df: pd.DataFrame, pdf_bytes: bytes):
    """
    Devuelve:
      - dict { (MODELO, PATRON): set(páginas) }
      - número total de páginas

    Lógica:
      - Buscamos los (MODELO, PATRON) dentro del texto de cada página.
      - Si una página NO contiene ningún modelo/patrón, se asigna a los mismos
        (MODELO, PATRON) que la última página donde sí aparecieron.
    """
    import pdfplumber

    if df.empty:
        return {}, 0

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        page_texts = [page.extract_text() or "" for page in pdf.pages]

    # Conjunto único de pares (MODELO, PATRON)
    keys = set()
    for _, row in df.iterrows():
        modelo = str(row.get("MODELO") or "").strip()
        patron = str(row.get("PATRON") or "").strip()
        if not modelo:
            continue
        keys.add((modelo, patron))

    model_map: dict[tuple[str, str], set[int]] = {k: set() for k in keys}

    last_keys_with_models: set[tuple[str, str]] = set()

    for page_num, text in enumerate(page_texts, start=1):
        current_keys_with_models: set[tuple[str, str]] = set()

        # Buscamos qué modelos/patrones aparecen en esta página
        for modelo, patron in keys:
            if not modelo:
                continue

            if modelo in text and (not patron or patron in text):
                model_map[(modelo, patron)].add(page_num)
                current_keys_with_models.add((modelo, patron))

        if current_keys_with_models:
            # Esta página tiene líneas con modelo → recordamos estos como "últimos vistos"
            last_keys_with_models = current_keys_with_models
        else:
            # Página sin modelos: la añadimos a los mismos modelos de la página anterior
            for k in last_keys_with_models:
                model_map[k].add(page_num)

    return model_map, len(page_texts)


def build_eci_model_page_map(df: pd.DataFrame, pdf_bytes: bytes):
    """
    Devuelve:
      - dict { MODELO: set(páginas) }
      - número total de páginas

    Lógica:
      - Buscamos el código MODELO dentro del texto de cada página.
      - Si una página NO contiene ningún modelo, se asigna a los mismos
        modelos que la última página donde sí aparecieron.
    """
    import pdfplumber

    if df.empty:
        return {}, 0

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        page_texts = [page.extract_text() or "" for page in pdf.pages]

    modelos = [
        str(m).strip()
        for m in df["MODELO"].dropna().unique()
        if str(m).strip()
    ]

    model_map: dict[str, set[int]] = {m: set() for m in modelos}
    last_models_with_data: set[str] = set()

    for page_num, text in enumerate(page_texts, start=1):
        current_models_with_data: set[str] = set()

        # ¿Qué modelos aparecen en esta página?
        for modelo in modelos:
            if modelo and modelo in text:
                model_map[modelo].add(page_num)
                current_models_with_data.add(modelo)

        if current_models_with_data:
            last_models_with_data = current_models_with_data
        else:
            # Página sin modelos → la añadimos a los últimos modelos vistos
            for m in last_models_with_data:
                model_map[m].add(page_num)

    return model_map, len(page_texts)


def create_eurofiel_folders_and_pdfs(df: pd.DataFrame, pdf_bytes: bytes, base_dir: str):
    """
    Crea una carpeta por MODELO+PATRON en Eurofiel y guarda dentro:
      - Un PDF con las páginas correspondientes a ese modelo/patrón.
      - Un Excel con el resumen filtrado a ese MODELO+PATRON,
        con bordes, cabeceras amarillas y fila TOTAL (pedidos / unidades).
    base_dir local demo: "./data/ediwin/out/eurofiel"
    """
    os.makedirs(base_dir, exist_ok=True)

    model_map, total_pages = build_eurofiel_model_page_map(df, pdf_bytes)
    if not model_map:
        return

    reader = PdfReader(BytesIO(pdf_bytes))

    for (modelo, patron), pages in model_map.items():
        if not pages:
            continue

        # Nombre carpeta: MODELO_PATRON saneado
        folder_name = f"{modelo}_{patron}"
        folder_name = safe_fs_name(folder_name)
        folder_path = os.path.join(base_dir, folder_name)
        os.makedirs(folder_path, exist_ok=True)

        # ---------- PDF por modelo/patrón ----------
        writer = PdfWriter()
        for p in sorted(pages):
            if 1 <= p <= total_pages:
                writer.add_page(reader.pages[p - 1])

        pdf_out_path = os.path.join(folder_path, f"{folder_name}.pdf")
        with open(pdf_out_path, "wb") as f_out:
            writer.write(f_out)

        # ---------- Excel filtrado por modelo/patrón ----------
        sub_df = df[
            (df["MODELO"].astype(str).str.strip() == str(modelo).strip())
            & (df["PATRON"].astype(str).str.strip() == str(patron).strip())
        ].copy()

        if sub_df.empty:
            continue

        # Aseguramos TOTAL_UNIDADES numérico
        if "TOTAL_UNIDADES" in sub_df.columns:
            sub_df["TOTAL_UNIDADES"] = pd.to_numeric(
                sub_df["TOTAL_UNIDADES"], errors="coerce"
            ).fillna(0).astype(int)
            total_unidades = int(sub_df["TOTAL_UNIDADES"].sum())
        else:
            total_unidades = None

        # Aseguramos tallas numéricas
        for t in TALLAS:
            if t in sub_df.columns:
                sub_df[t] = pd.to_numeric(sub_df[t], errors="coerce").fillna(0).astype(int)

        # Detectamos tallas activas SOLO en este modelo
        active_sizes_model = []
        for t in TALLAS:
            if t in sub_df.columns and sub_df[t].sum() > 0:
                active_sizes_model.append(t)

        # Eliminamos columnas de tallas que estén todo a 0
        zero_sizes_model = [t for t in TALLAS if t in sub_df.columns and t not in active_sizes_model]
        if zero_sizes_model:
            sub_df = sub_df.drop(columns=zero_sizes_model)

        # Total de pedidos (columna PEDIDO)
        total_pedidos = None
        if "PEDIDO" in sub_df.columns:
            total_pedidos = sub_df["PEDIDO"].nunique()

        # Fila TOTAL al final
        cols = list(sub_df.columns)
        total_row = {col: "" for col in cols}
        if cols:
            total_row[cols[0]] = "TOTAL"
        if "PEDIDO" in cols and total_pedidos is not None:
            total_row["PEDIDO"] = total_pedidos
        if "TOTAL_UNIDADES" in cols and total_unidades is not None:
            total_row["TOTAL_UNIDADES"] = total_unidades

        # Totales por talla (solo las que quedan)
        for t in active_sizes_model:
            if t in cols:
                total_row[t] = int(sub_df[t].sum())


        sub_df_with_total = pd.concat(
            [sub_df, pd.DataFrame([total_row])], ignore_index=True
        )

        excel_out_path = os.path.join(folder_path, f"{folder_name}.xlsx")
        with pd.ExcelWriter(excel_out_path, engine="openpyxl") as writer:
            sub_df_with_total.to_excel(writer, index=False, sheet_name="Resumen")
            wb = writer.book
            style_workbook_with_borders_and_headers(wb)



def create_eci_folders_and_pdfs(df: pd.DataFrame, pdf_bytes: bytes, base_dir: str):
    """
    Crea una carpeta por MODELO en ECI y guarda dentro:
      - Un PDF con las páginas correspondientes a ese modelo.
      - Un Excel con el resumen filtrado a ese MODELO,
        con bordes, cabeceras amarillas y fila TOTAL (pedidos / unidades).
    base_dir local demo: "./data/ediwin/out/eci"
    """
    os.makedirs(base_dir, exist_ok=True)

    model_map, total_pages = build_eci_model_page_map(df, pdf_bytes)
    if not model_map:
        return

    reader = PdfReader(BytesIO(pdf_bytes))

    for modelo, pages in model_map.items():
        if not pages:
            continue

        # Nombre carpeta: MODELO saneado
        folder_name = safe_fs_name(modelo)
        folder_path = os.path.join(base_dir, folder_name)
        os.makedirs(folder_path, exist_ok=True)

        # ---------- PDF por modelo ----------
        writer = PdfWriter()
        for p in sorted(pages):
            if 1 <= p <= total_pages:
                writer.add_page(reader.pages[p - 1])

        pdf_out_path = os.path.join(folder_path, f"{folder_name}.pdf")
        with open(pdf_out_path, "wb") as f_out:
            writer.write(f_out)

        # ---------- Excel filtrado por modelo ----------
        sub_df = df[df["MODELO"].astype(str).str.strip() == str(modelo).strip()].copy()
        if sub_df.empty:
            continue

        # Aseguramos TOTAL_UNIDADES numérico
        if "TOTAL_UNIDADES" in sub_df.columns:
            sub_df["TOTAL_UNIDADES"] = pd.to_numeric(
                sub_df["TOTAL_UNIDADES"], errors="coerce"
            ).fillna(0).astype(int)
            total_unidades = int(sub_df["TOTAL_UNIDADES"].sum())
        else:
            total_unidades = None

        # Aseguramos tallas numéricas
        for t in TALLAS:
            if t in sub_df.columns:
                sub_df[t] = pd.to_numeric(sub_df[t], errors="coerce").fillna(0).astype(int)

        # Detectamos tallas activas SOLO en este modelo
        active_sizes_model = []
        for t in TALLAS:
            if t in sub_df.columns and sub_df[t].sum() > 0:
                active_sizes_model.append(t)

        # Eliminamos columnas de tallas que estén todo a 0
        zero_sizes_model = [t for t in TALLAS if t in sub_df.columns and t not in active_sizes_model]
        if zero_sizes_model:
            sub_df = sub_df.drop(columns=zero_sizes_model)

        # Total de pedidos (columna N_PEDIDO)
        total_pedidos = None
        if "N_PEDIDO" in sub_df.columns:
            total_pedidos = sub_df["N_PEDIDO"].nunique()

        # Fila TOTAL al final
        cols = list(sub_df.columns)
        total_row = {col: "" for col in cols}
        if cols:
            total_row[cols[0]] = "TOTAL"
        if "N_PEDIDO" in cols and total_pedidos is not None:
            total_row["N_PEDIDO"] = total_pedidos
        if "TOTAL_UNIDADES" in cols and total_unidades is not None:
            total_row["TOTAL_UNIDADES"] = total_unidades

        # Totales por talla (solo las que quedan en este modelo)
        for t in active_sizes_model:
            if t in cols:
                total_row[t] = int(sub_df[t].sum())

        sub_df_with_total = pd.concat(
            [sub_df, pd.DataFrame([total_row])], ignore_index=True
        )

        excel_out_path = os.path.join(folder_path, f"{folder_name}.xlsx")
        with pd.ExcelWriter(excel_out_path, engine="openpyxl") as writer:
            sub_df_with_total.to_excel(writer, index=False, sheet_name="Resumen")
            wb = writer.book
            style_workbook_with_borders_and_headers(wb)

def split_ediwin_txt_files_per_model_eurofiel(
    base_dir: str,
    df: pd.DataFrame,                 # <- no se usa pero lo mantenemos por compatibilidad
    linped_bytes: bytes,
    linped_name: str,
    cabped_bytes: bytes | None = None,
    cabped_name: str | None = None,
    locped_bytes: bytes | None = None,
    locped_name: str | None = None,
    obsped_bytes: bytes | None = None,
    obsped_name: str | None = None,
    obslped_bytes: bytes | None = None,
    obslped_name: str | None = None,
    encoding: str = "latin-1",
    max_model_length: int | None = None,   # 👈 aquí sí va el límite
):
    """
    EUROFIEL:
    Reparte los TXT de EDIWIN por carpeta {MODELO}_{PATRON} usando SOLO la info de LINPED.

    - Del LINPED cogemos:
        pedido_int = campo 0  → '00135142'
        cod_cli    = campo 5  → '7603709/19/01'  → PATRON   = '7603709/19'
        cod_prov   = campo 6  → '3VS596 EST/MULTICOLOR/XS' → MODELO = '3VS596 EST/MULTICOLOR'
    - Carpeta = safe_fs_name("{MODELO}_{PATRON}") → '3VS596-EST-MULTICOLOR_7603709-19'
    - Luego, cualquier línea de CABPED/LOCPED/OBSPED/OBSLPED con ese nº de pedido
      va a parar a esa(s) carpeta(s).

    - Si max_model_length no es None, recortamos el campo cod_prov (MODELO)
      en las líneas LINPED antes de guardar los TXT, para cumplir con Sage.
    """

    if linped_bytes is None:
        raise ValueError("LINPED es obligatorio para poder repartir los TXT por modelo.")

    from collections import defaultdict

    def decode_lines(b: bytes | None) -> list[str]:
        if b is None:
            return []
        text = b.decode(encoding, errors="ignore")
        return text.splitlines(keepends=True)

    # --- 1) Leemos todas las líneas de los TXT ---
    lin_lines = decode_lines(linped_bytes)
    cab_lines = decode_lines(cabped_bytes)
    loc_lines = decode_lines(locped_bytes)
    obs_lines = decode_lines(obsped_bytes)
    obsl_lines = decode_lines(obslped_bytes)

    # --- Helper: recortar el campo MODELO (cod_prov) en LINPED para Sage ---
    # 🔎 aquí guardaremos qué modelo se ha recortado a qué
    model_changes: dict[str, str] = {}

    def shorten_cod_prov_for_sage(cod_prov: str, max_len: int) -> str:
        """
        Ejemplo de entrada: '3VS596 EST/MULTICOLOR/XS'
        - Separa talla (último /)
        - Recorta la parte izquierda (modelo+color) a max_len caracteres
        - Devuelve algo tipo '3VS596 EST/MULTICO/XS'
        """
        cod_prov = cod_prov.strip()
        if max_len is None or max_len <= 0:
            return cod_prov

        # Caso típico: MODELO/COLOR/TALLA
        if "/" in cod_prov:
            base, talla = cod_prov.rsplit("/", 1)
            base = base.strip()
            original_base = base

            if len(base) > max_len:
                new_base = base[:max_len]
            else:
                new_base = base

            # Registramos el cambio SOLO si ha habido recorte
            if new_base != original_base:
                model_changes.setdefault(original_base, new_base)

            return f"{new_base}/{talla}"

        # Por si viniera sin talla
        original_base = cod_prov
        if len(cod_prov) > max_len:
            new_base = cod_prov[:max_len]
            if new_base != original_base:
                model_changes.setdefault(original_base, new_base)
            return new_base

        return cod_prov


    def adjust_lin_line_model(line: str) -> str:
        """
        Si max_model_length está definido, recorta el campo cod_prov (índice 6)
        de la línea LINPED. Respeta saltos de línea originales.
        """
        if max_model_length is None:
            return line
        if not line.strip():
            return line

        newline = ""
        if line.endswith("\r\n"):
            newline = "\r\n"
        elif line.endswith("\n"):
            newline = "\n"

        core = line.rstrip("\r\n")
        parts = core.split(";")
        if len(parts) < 7:
            return line  # no tiene la estructura esperada, no tocamos

        cod_prov = parts[6].strip()
        if not cod_prov:
            return line

        parts[6] = shorten_cod_prov_for_sage(cod_prov, max_model_length)
        return ";".join(parts) + newline

    # --- 2) Mapa pedido_int -> set de carpetas {MODELO}_{PATRON} usando SOLO LINPED ---
    pedido_to_folders: dict[str, set[str]] = defaultdict(set)

    for line in lin_lines:
        if not line.strip():
            continue

        parts = line.split(";")
        if len(parts) < 7:
            continue

        pedido_int = parts[0].strip()
        cod_cli = parts[5].strip()   # 7603709/19/01
        cod_prov = parts[6].strip()  # 3VS596 EST/MULTICOLOR/XS

        if not pedido_int or not cod_cli or not cod_prov:
            continue

        # PATRON: quitamos la talla → 7603709/19
        patron = re.sub(r"/[^/]+$", "", cod_cli)
        # MODELO: quitamos la talla → 3VS596 EST/MULTICOLOR
        modelo = re.sub(r"/[^/]+$", "", cod_prov)

        folder_name = safe_fs_name(f"{modelo}_{patron}")
        pedido_to_folders[pedido_int].add(folder_name)

    if not pedido_to_folders:
        return

    # --- 3) Dado un listado de líneas, las repartimos por carpeta usando pedido_int ---
    def distribute_by_folder(lines: list[str]) -> dict[str, list[str]]:
        out: dict[str, list[str]] = defaultdict(list)
        for line in lines:
            if not line.strip():
                continue
            parts = line.split(";")
            if not parts:
                continue
            pedido_int = parts[0].strip()
            if not pedido_int:
                continue

            folders = pedido_to_folders.get(pedido_int)
            if not folders:
                continue

            for folder in folders:
                out[folder].append(line)

        return out

    lin_by_folder = distribute_by_folder(lin_lines)
    cab_by_folder = distribute_by_folder(cab_lines)
    loc_by_folder = distribute_by_folder(loc_lines)
    obs_by_folder = distribute_by_folder(obs_lines)
    obsl_by_folder = distribute_by_folder(obsl_lines)

    # --- 4) Lista final de carpetas a tocar ---
    all_folders = sorted(
        {folder for folders in pedido_to_folders.values() for folder in folders}
    )

    # --- 5) Escribimos los TXT dentro de cada carpeta {MODELO}_{PATRON} ---
    os.makedirs(base_dir, exist_ok=True)

    for folder in all_folders:
        folder_path = os.path.join(base_dir, folder)
        os.makedirs(folder_path, exist_ok=True)

        # LINPED (obligatorio)
        if linped_name:
            lin_path = os.path.join(folder_path, linped_name)
            with open(lin_path, "w", encoding=encoding, newline="") as f:
                raw_lines = lin_by_folder.get(folder, [])
                adj_lines = [adjust_lin_line_model(line) for line in raw_lines]
                f.writelines(adj_lines)

        # CABPED
        if cabped_name:
            cab_path = os.path.join(folder_path, cabped_name)
            with open(cab_path, "w", encoding=encoding, newline="") as f:
                f.writelines(cab_by_folder.get(folder, []))

        # LOCPED
        if locped_name:
            loc_path = os.path.join(folder_path, locped_name)
            with open(loc_path, "w", encoding=encoding, newline="") as f:
                f.writelines(loc_by_folder.get(folder, []))

        # OBSPED
        if obsped_name:
            obs_path = os.path.join(folder_path, obsped_name)
            with open(obs_path, "w", encoding=encoding, newline="") as f:
                f.writelines(obs_by_folder.get(folder, []))

        # OBSLPED
        if obslped_name:
            obsl_path = os.path.join(folder_path, obslped_name)
            with open(obsl_path, "w", encoding=encoding, newline="") as f:
                f.writelines(obsl_by_folder.get(folder, []))

    return model_changes

def split_ediwin_txt_files_per_model(
    base_dir: str,
    linped_bytes: bytes,
    linped_name: str,
    cabped_bytes: bytes | None = None,
    cabped_name: str | None = None,
    locped_bytes: bytes | None = None,
    locped_name: str | None = None,
    obsped_bytes: bytes | None = None,
    obsped_name: str | None = None,
    obslped_bytes: bytes | None = None,
    obslped_name: str | None = None,
    encoding: str = "latin-1",
):
    """
    ECI:
    Reparte los TXT de EDIWIN por carpeta MODELO usando SOLO la info de LINPED.

    - Del LINPED cogemos:
        pedido_int = campo 0
        cod_prov   = campo 6  → suele llevar MODELO/COLOR/TALLA o similar
    - Carpeta = safe_fs_name(MODELO_sin_talla)
    - Luego, cualquier línea de CABPED/LOCPED/OBSPED/OBSLPED con ese nº de pedido
      va a parar a esa(s) carpeta(s).
    """

    if linped_bytes is None:
        raise ValueError("LINPED es obligatorio para poder repartir los TXT por modelo.")

    def decode_lines(b: bytes | None) -> list[str]:
        if b is None:
            return []
        text = b.decode(encoding, errors="ignore")
        return text.splitlines(keepends=True)

    # --- 1) Leemos todas las líneas ---
    lin_lines = decode_lines(linped_bytes)
    cab_lines = decode_lines(cabped_bytes)
    loc_lines = decode_lines(locped_bytes)
    obs_lines = decode_lines(obsped_bytes)
    obsl_lines = decode_lines(obslped_bytes)

    from collections import defaultdict
    pedido_to_folders: dict[str, set[str]] = defaultdict(set)

    # --- 2) Mapa pedido_int -> set de carpetas (MODELO) usando SOLO LINPED ---
    for line in lin_lines:
        if not line.strip():
            continue

        parts = line.split(";")
        if len(parts) < 7:
            continue

        pedido_int = parts[0].strip()
        cod_prov = parts[6].strip()

        if not pedido_int or not cod_prov:
            continue

        # MODELO: quitamos la talla (último /) si existe
        modelo = re.sub(r"/[^/]+$", "", cod_prov).strip()

        folder_name = safe_fs_name(modelo or "SIN_MODELO")
        pedido_to_folders[pedido_int].add(folder_name)

    if not pedido_to_folders:
        return

    # --- 3) Dado un listado de líneas, las repartimos por carpeta usando pedido_int ---
    def distribute_by_folder(lines: list[str]) -> dict[str, list[str]]:
        out: dict[str, list[str]] = defaultdict(list)
        for line in lines:
            if not line.strip():
                continue
            parts = line.split(";")
            if not parts:
                continue
            pedido_int = parts[0].strip()
            if not pedido_int:
                continue

            folders = pedido_to_folders.get(pedido_int)
            if not folders:
                continue

            for folder in folders:
                out[folder].append(line)

        return out

    lin_by_folder = distribute_by_folder(lin_lines)
    cab_by_folder = distribute_by_folder(cab_lines)
    loc_by_folder = distribute_by_folder(loc_lines)
    obs_by_folder = distribute_by_folder(obs_lines)
    obsl_by_folder = distribute_by_folder(obsl_lines)

    # --- 4) Lista final de carpetas a tocar ---
    all_folders = sorted(
        {folder for folders in pedido_to_folders.values() for folder in folders}
    )

    # --- 5) Escribimos los TXT dentro de cada carpeta MODELO ---
    os.makedirs(base_dir, exist_ok=True)

    for folder in all_folders:
        folder_path = os.path.join(base_dir, folder)
        os.makedirs(folder_path, exist_ok=True)

        # LINPED
        if linped_name:
            lin_path = os.path.join(folder_path, linped_name)
            with open(lin_path, "w", encoding=encoding, newline="") as f:
                f.writelines(lin_by_folder.get(folder, []))

        # CABPED
        if cabped_name:
            cab_path = os.path.join(folder_path, cabped_name)
            with open(cab_path, "w", encoding=encoding, newline="") as f:
                f.writelines(cab_by_folder.get(folder, []))

        # LOCPED
        if locped_name:
            loc_path = os.path.join(folder_path, locped_name)
            with open(loc_path, "w", encoding=encoding, newline="") as f:
                f.writelines(loc_by_folder.get(folder, []))

        # OBSPED
        if obsped_name:
            obs_path = os.path.join(folder_path, obsped_name)
            with open(obs_path, "w", encoding=encoding, newline="") as f:
                f.writelines(obs_by_folder.get(folder, []))

        # OBSLPED
        if obslped_name:
            obsl_path = os.path.join(folder_path, obslped_name)
            with open(obsl_path, "w", encoding=encoding, newline="") as f:
                f.writelines(obsl_by_folder.get(folder, []))


# ============= STREAMLIT APP =============

st.set_page_config(page_title="Resumen pedidos EDIWIN", layout="wide")

st.title("📦 Resumen de pedidos desde PDF EDIWIN")

cliente = st.selectbox("Cliente", ["Eurofiel", "El Corte Inglés"])

st.write(
    "Sube un PDF de **Eurofiel** o **El Corte Inglés** y te saco "
    "las líneas clave listas para Excel."
)

label = "📁 Sube tu PDF de Eurofiel" if cliente == "Eurofiel" else "📁 Sube tu PDF de El Corte Inglés"
uploaded_pdf = st.file_uploader(label, type=["pdf"])

if uploaded_pdf is not None:
    try:
        if cliente == "Eurofiel":
            df = parse_pdf_eurofiel_bytes(uploaded_pdf.getvalue())
        else:
            df = parse_pdf_eci_bytes(uploaded_pdf.getvalue())

        if df.empty:
            st.warning("No se han detectado pedidos en el PDF. Revisa el formato.")
        else:
            st.subheader("📊 Vista previa de pedidos detectados")

            # Aseguramos TOTAL_UNIDADES numérico
            if "TOTAL_UNIDADES" in df.columns:
                df["TOTAL_UNIDADES"] = pd.to_numeric(
                    df["TOTAL_UNIDADES"], errors="coerce"
                ).fillna(0)

            # ===== Limpiamos tallas que están todo a 0 =====
            active_sizes = get_active_sizes(df)
            zero_size_cols = [t for t in TALLAS if t in df.columns and t not in active_sizes]
            df_display = df.drop(columns=zero_size_cols)

            # Estilos por MODELO usando solo tallas activas
            styled_df = style_by_model(df_display)
            st.dataframe(styled_df, use_container_width=True)

            # ====== RESÚMENES Y EXPORT ======
            if cliente == "Eurofiel":
                st.subheader("📦 Resumen por MODELO")

                resumen = (
                    df.groupby("MODELO", dropna=False)
                    .agg(
                        PEDIDOS=("PEDIDO", "nunique"),
                        UNIDADES_TOTALES=("TOTAL_UNIDADES", "sum"),
                    )
                    .reset_index()
                    .sort_values("PEDIDOS", ascending=False)
                )

                st.dataframe(resumen, use_container_width=True)

                # --- versión para Excel con fila TOTAL ---
                resumen_xlsx = resumen.copy()
                total_row = {
                    "MODELO": "TOTAL",
                    "PEDIDOS": resumen_xlsx["PEDIDOS"].sum(),
                    "UNIDADES_TOTALES": resumen_xlsx["UNIDADES_TOTALES"].sum(),
                }
                resumen_xlsx = pd.concat(
                    [resumen_xlsx, pd.DataFrame([total_row])],
                    ignore_index=True,
                )

                # ---- Exportar Excel Eurofiel ----
                excel_buffer = BytesIO()
                with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                    # Ojo: styled_df ya está hecho sobre df_display
                    styled_df.to_excel(writer, sheet_name="Pedidos", index=False)
                    resumen_xlsx.to_excel(writer, sheet_name="Resumen por modelo", index=False)

                    wb = writer.book
                    style_workbook_with_borders_and_headers(wb)

                st.download_button(
                    label="⬇️ Descargar Excel",
                    data=excel_buffer.getvalue(),
                    file_name="eurofiel_resumen_pedidos.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                # CSV simple (también sin tallas 0)
                csv_bytes = df_display.to_csv(index=False).encode("utf-8")
                st.download_button(
                    label="⬇️ Descargar CSV",
                    data=csv_bytes,
                    file_name="eurofiel_resumen_pedidos.csv",
                    mime="text/csv",
                )

                # ---- Crear carpetas y PDFs por modelo en Eurofiel ----
                if st.button("📂 Crear carpetas y PDFs por modelo en Eurofiel"):
                    try:
                        EUROFIEL_BASE_DIR = r"./data/ediwin/out/eurofiel"
                        create_eurofiel_folders_and_pdfs(
                            df, uploaded_pdf.getvalue(), EUROFIEL_BASE_DIR
                        )
                        st.success(
                            f"Carpetas y PDFs creados en '{EUROFIEL_BASE_DIR}'."
                        )
                    except Exception as e:
                        st.error(f"Error creando carpetas/PDFs: {e}")

                st.markdown("### TXT EDIWIN → repartir por modelo (EUROFIEL)")

                edi_files = st.file_uploader(
                    "Sube los TXT de EDIWIN (CABPED, LINPED, LOCPED, OBSPED, OBSLPED)",
                    type=["txt", "TXT"],
                    accept_multiple_files=True,
                    key="ediwin_txt_eurofiel",
                )

                # ✅ Opción para recortar MODELO (cod_prov LINPED) a 20 caracteres
                recortar_modelo = st.checkbox(
                    "Recortar campo MODELO de LINPED a 20 caracteres para Sage",
                    value=True,
                    help="Si lo activas, el programa acorta el MODELO (cod_prov) en LINPED para que Sage no lo rechace."
                )

                if edi_files:
                    cab_file = lin_file = loc_file = obs_file = obsl_file = None

                    for f in edi_files:
                        name_upper = f.name.upper()
                        if "LINPED" in name_upper:
                            lin_file = f
                        elif "CABPED" in name_upper:
                            cab_file = f
                        elif "LOCPED" in name_upper:
                            loc_file = f
                        elif "OBSPED" in name_upper:
                            obs_file = f
                        elif "OBSLPED" in name_upper:
                            obsl_file = f

                    if st.button("📂 Repartir TXT EDIWIN por modelo en carpetas EUROFIEL"):
                        if lin_file is None:
                            st.error("Falta el fichero LINPED_*.TXT, es obligatorio para repartir por modelo.")
                        else:
                            try:
                                max_len = 20 if recortar_modelo else None

                                model_changes = split_ediwin_txt_files_per_model_eurofiel(
                                    base_dir=EUROFIEL_BASE_DIR,
                                    df=df,
                                    linped_bytes=lin_file.getvalue(),
                                    linped_name=lin_file.name,
                                    cabped_bytes=cab_file.getvalue() if cab_file else None,
                                    cabped_name=cab_file.name if cab_file else None,
                                    locped_bytes=loc_file.getvalue() if loc_file else None,
                                    locped_name=loc_file.name if loc_file else None,
                                    obsped_bytes=obs_file.getvalue() if obs_file else None,
                                    obsped_name=obs_file.name if obs_file else None,
                                    obslped_bytes=obsl_file.getvalue() if obsl_file else None,
                                    obslped_name=obsl_file.name if obsl_file else None,
                                    max_model_length=max_len,  # 👈 importante
                                )

                                st.success("TXT repartidos por modelo y copiados en sus carpetas EUROFIEL.")

                                # 👀 Visualización de los modelos recortados para SAGE
                                if max_len is not None:
                                    if model_changes:
                                        st.markdown("#### Modelos ajustados para SAGE (nombre corto)")

                                        df_models = pd.DataFrame(
                                            [
                                                {"MODELO_ORIGINAL": k, "MODELO_SAGE": v}
                                                for k, v in sorted(model_changes.items())
                                            ]
                                        )
                                        st.dataframe(df_models, use_container_width=True)
                                    else:
                                        st.info(
                                            "No se ha recortado ningún modelo: todos caben en "
                                            f"{max_len} caracteres."
                                        )

                            except Exception as e:
                                st.error(f"Error al repartir TXT: {e}")




            else:  # El Corte Inglés
                st.subheader("📦 Resumen por MODELO + COLOR")

                resumen_mc = (
                    df.groupby(["MODELO", "COLOR"], dropna=False)
                    .agg(
                        PEDIDOS=("N_PEDIDO", "nunique"),
                        UNIDADES_TOTALES=("TOTAL_UNIDADES", "sum"),
                    )
                    .reset_index()
                    .sort_values("PEDIDOS", ascending=False)
                )

                st.dataframe(resumen_mc, use_container_width=True)

                st.subheader("🧩 Resumen por MODELO (todas las sucursales/colores)")

                resumen_m = (
                    df.groupby("MODELO", dropna=False)
                    .agg(
                        PEDIDOS=("N_PEDIDO", "nunique"),
                        UNIDADES_TOTALES=("TOTAL_UNIDADES", "sum"),
                    )
                    .reset_index()
                    .sort_values("PEDIDOS", ascending=False)
                )

                st.dataframe(resumen_m, use_container_width=True)

                # --- versiones para Excel con fila TOTAL ---
                resumen_mc_xlsx = resumen_mc.copy()
                total_mc = {
                    "MODELO": "TOTAL",
                    "COLOR": "",
                    "PEDIDOS": resumen_mc_xlsx["PEDIDOS"].sum(),
                    "UNIDADES_TOTALES": resumen_mc_xlsx["UNIDADES_TOTALES"].sum(),
                }
                resumen_mc_xlsx = pd.concat(
                    [resumen_mc_xlsx, pd.DataFrame([total_mc])],
                    ignore_index=True,
                )

                resumen_m_xlsx = resumen_m.copy()
                total_m = {
                    "MODELO": "TOTAL",
                    "PEDIDOS": resumen_m_xlsx["PEDIDOS"].sum(),
                    "UNIDADES_TOTALES": resumen_m_xlsx["UNIDADES_TOTALES"].sum(),
                }
                resumen_m_xlsx = pd.concat(
                    [resumen_m_xlsx, pd.DataFrame([total_m])],
                    ignore_index=True,
                )

                # ---- Exportar Excel ECI ----
                excel_buffer = BytesIO()
                with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                    styled_df.to_excel(writer, sheet_name="Pedidos", index=False)
                    resumen_mc_xlsx.to_excel(writer, sheet_name="Resumen modelo+color", index=False)
                    resumen_m_xlsx.to_excel(writer, sheet_name="Resumen modelo", index=False)

                    wb = writer.book
                    style_workbook_with_borders_and_headers(wb)
                st.download_button(
                    label="⬇️ Descargar Excel",
                    data=excel_buffer.getvalue(),
                    file_name="eci_resumen_pedidos.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                # CSV simple
                csv_bytes = df_display.to_csv(index=False).encode("utf-8")
                st.download_button(
                    label="⬇️ Descargar CSV",
                    data=csv_bytes,
                    file_name="eci_resumen_pedidos.csv",
                    mime="text/csv",
                )

                # ---- Crear carpetas y PDFs por modelo en ECI ----
                if st.button("📂 Crear carpetas y PDFs por modelo en ECI"):
                    try:
                        ECI_BASE_DIR = r"./data/ediwin/out/eci"
                        create_eci_folders_and_pdfs(
                            df, uploaded_pdf.getvalue(), ECI_BASE_DIR
                        )
                        st.success(
                            f"Carpetas y PDFs creados en '{ECI_BASE_DIR}'."
                        )
                    except Exception as e:
                        st.error(f"Error creando carpetas/PDFs: {e}")

                st.markdown("### TXT EDIWIN → repartir por modelo")

                edi_files = st.file_uploader(
                    "Sube los TXT de EDIWIN (CABPED, LINPED, LOCPED, OBSPED, OBSLPED)",
                    type=["txt", "TXT"],
                    accept_multiple_files=True,
                )

                if edi_files:
                    # Clasificamos por nombre
                    cab_file = lin_file = loc_file = obs_file = obsl_file = None

                    for f in edi_files:
                        name_upper = f.name.upper()
                        if "LINPED" in name_upper:
                            lin_file = f
                        elif "CABPED" in name_upper:
                            cab_file = f
                        elif "LOCPED" in name_upper:
                            loc_file = f
                        elif "OBSLPED" in name_upper:
                            obsl_file = f
                        elif "OBSPED" in name_upper:
                            obs_file = f

                    if st.button("📂 Repartir TXT EDIWIN por modelo en carpetas ECI"):
                        if lin_file is None:
                            st.error(
                                "Falta el fichero LINPED_*.TXT, es obligatorio para repartir por modelo."
                            )
                        else:
                            try:
                                ECI_BASE_DIR = r"./data/ediwin/out/eci"

                                split_ediwin_txt_files_per_model(
                                    base_dir=ECI_BASE_DIR,
                                    linped_bytes=lin_file.getvalue(),
                                    linped_name=lin_file.name,
                                    cabped_bytes=cab_file.getvalue() if cab_file else None,
                                    cabped_name=cab_file.name if cab_file else None,
                                    locped_bytes=loc_file.getvalue() if loc_file else None,
                                    locped_name=loc_file.name if loc_file else None,
                                    obsped_bytes=obs_file.getvalue() if obs_file else None,
                                    obsped_name=obs_file.name if obs_file else None,
                                    obslped_bytes=obsl_file.getvalue() if obsl_file else None,
                                    obslped_name=obsl_file.name if obsl_file else None,
                                )

                                st.success(
                                    "TXT repartidos por modelo y copiados en sus carpetas ECI."
                                )
                            except Exception as e:
                                st.error(f"Error al repartir TXT: {e}")



    except Exception as e:
        st.error(f"❌ Error procesando el PDF: {e}")
else:
    st.info("Sube un PDF para empezar.")


# ===== FOOTER =====
st.markdown("""
<hr style="margin-top: 60px;">

<div style="text-align:center; font-size:14px; color:#777;">
    Creado y desarrollado con mucho amor ❤️ por<br>
    <strong style="font-size:16px;">Equipo Demo</strong><br>
    <span style="font-size:12px;">@elvasco.x</span>
</div>
""", unsafe_allow_html=True)
