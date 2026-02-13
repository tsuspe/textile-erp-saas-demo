# app/(app)/tools/almacen/globalia-stock/cli.py
"""
CLI wrapper para integrar el Gestor de Stock (Python) dentro de Next.js.

Objetivo: CLONAR Streamlit (st_app.py) en Next manteniendo:
- mismas funcionalidades,
- misma lógica/cálculos,
- mismas exportaciones/importaciones,
- mismos saneos,
- mismos backups.

Convención:
- Para previews/listados: imprime JSON por stdout {ok:true, ...}
- Para exports: genera un ZIP en --out y devuelve {ok:true, out:"..."} por stdout.
"""

#!/usr/bin/env python3
from __future__ import annotations

import sys
from pathlib import Path

# Asegurar imports locales (backend y carpeta actual) aunque el cwd cambie
HERE = Path(__file__).resolve().parent
BACKEND_DIR = HERE / "backend"
sys.path.insert(0, str(BACKEND_DIR))
sys.path.insert(0, str(HERE))

import argparse
import io
import json
import os
import shutil
import sys
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import contextlib


import pandas as pd

# Importar el core (misma carpeta)
from gestor_oop import GestorStock, norm_codigo, norm_talla, parse_fecha_excel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# -----------------------
# JSON helpers (stdout)
# -----------------------
def _ok(**payload):
    print(json.dumps({"ok": True, **payload}, ensure_ascii=False))
    return 0


def _fail(code: str, detail: str = ""):
    print(
        json.dumps({"ok": False, "error": code, "detail": detail}, ensure_ascii=False)
    )
    return 1


def _read_env_path(key: str, fallback: str = "") -> str:
    v = os.environ.get(key, "") or ""
    return v.strip() or fallback


def _timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def _make_mgr(args) -> GestorStock:
    return GestorStock(
        path_inventario=args.inv,
        path_prevision=args.prev,
        path_talleres=args.talleres,
        path_clientes=args.clientes,
        export_dir=args.export_dir or None,
        backup_dir=args.backup_dir or None,
    )


def _capture_io(fn):
    """
    Captura stdout/stderr de llamadas al backend (gestor_oop),
    para que el CLI SOLO imprima JSON limpio.
    """
    out = io.StringIO()
    err = io.StringIO()
    with contextlib.redirect_stdout(out), contextlib.redirect_stderr(err):
        res = fn()
    return res, out.getvalue(), err.getvalue()


# -----------------------
# Excel formatting helpers (copiados de Streamlit)
# -----------------------
BRIGHT_YELLOW = "FFFF00"  # amarillo Excel fuerte
PASTEL_YELLOW = "FFF9C4"  # amarillo suave

_PAST_GREENS = [
    "#c8e6c9",
    "#a5d6a7",
    "#81c784",
    "#66bb6a",
    "#4caf50",
    "#2e7d32",
    "#1b5e20",
]
_FUTURE_REDS = [
    "#ffcdd2",
    "#ef9a9a",
    "#e57373",
    "#ef5350",
    "#d32f2f",
    "#b71c1c",
    "#7f0000",
]


def _month_index(d: date) -> int:
    return d.year * 12 + d.month


def _month_delta_color(delta: int) -> str:
    if delta == 0:
        return ""
    if delta < 0:
        idx = min(abs(delta) - 1, len(_PAST_GREENS) - 1)
        return _PAST_GREENS[idx]
    idx = min(delta - 1, len(_FUTURE_REDS) - 1)
    return _FUTURE_REDS[idx]


def _parse_date_flexible(val):
    if isinstance(val, date):
        return val
    if not val:
        return None
    s = str(val).strip()
    s10 = s[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s10, fmt).date()
        except Exception:
            pass
    return None


def _auto_qty_col(df, candidates=None):
    if candidates is None:
        candidates = [
            "STOCK",
            "stock",
            "STOCK_ESTIMADO",
            "stock_estimado",
            "ESTIMADO",
            "estimado",
            "QTY",
            "qty",
            "CANTIDAD",
            "cantidad",
            "TOTAL",
            "total",
        ]
    for c in candidates:
        if c in df.columns:
            return c
    blacklist = {
        "MODELO",
        "modelo",
        "TALLA",
        "talla",
        "PEDIDO",
        "pedido",
        "CLIENTE",
        "cliente",
        "COLOR",
        "color",
        "DESCRIPCION",
        "descripcion",
    }
    for col in df.columns:
        if col in blacklist:
            continue
        try:
            pd.to_numeric(df[col])
            return col
        except Exception:
            continue
    return None


def _excel_yellow_header_and_total(
    ws, header_row: int = 1, highlight_last: bool = True
):
    bright_fill = PatternFill(fill_type="solid", fgColor=BRIGHT_YELLOW)
    bold_font = Font(bold=True)
    for cell in ws[header_row]:
        cell.fill = bright_fill
        cell.font = bold_font
    if highlight_last and ws.max_row >= 2:
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.fill = bright_fill
            cell.font = bold_font


def _excel_add_borders(ws, min_row: int = 1, min_col: int = 1):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(
        min_row=min_row, max_row=ws.max_row, min_col=min_col, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = border


def _excel_highlight_totals_by_talla(
    ws, df: pd.DataFrame, talla_col: str = "TALLA", data_start_row: int = 2
):
    if talla_col not in df.columns:
        return
    try:
        talla_idx = list(df.columns).index(talla_col) + 1
    except ValueError:
        return

    bright_fill = PatternFill(fill_type="solid", fgColor=BRIGHT_YELLOW)
    bold_font = Font(bold=True)

    for row in range(data_start_row, ws.max_row + 1):
        val = ws.cell(row=row, column=talla_idx).value
        if not val:
            continue
        txt = str(val).strip().upper()
        if txt in ("TOTAL MODELO", "TOTAL GENERAL"):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = bright_fill
                cell.font = bold_font


def _excel_color_stock_ranges(
    ws, df: pd.DataFrame, qty_col_candidates=None, data_start_row: int = 2
):
    qty_col = _auto_qty_col(df, candidates=qty_col_candidates)
    if not qty_col:
        return
    try:
        col_idx = list(df.columns).index(qty_col) + 1
    except ValueError:
        return

    red_fill = PatternFill(fill_type="solid", fgColor="FFCDD2")
    orange_fill = PatternFill(fill_type="solid", fgColor="FFE0B2")
    yellow_fill = PatternFill(fill_type="solid", fgColor=PASTEL_YELLOW)

    for row in range(data_start_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        val = cell.value
        try:
            q = float(str(val).replace(",", "."))
        except Exception:
            continue

        fill = None
        if q <= 0:
            fill = red_fill
        elif 0 < q <= 10:
            fill = orange_fill
        elif 10 < q <= 25:
            fill = yellow_fill

        if fill:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=row, column=c).fill = fill


def _excel_color_pend_by_month(
    ws, df: pd.DataFrame, date_col: str = "FECHA", data_start_row: int = 2
):
    if date_col not in df.columns:
        return
    today = date.today()
    date_col_idx = list(df.columns).index(date_col) + 1

    for row in range(data_start_row, ws.max_row):
        cell = ws.cell(row=row, column=date_col_idx)
        d = _parse_date_flexible(cell.value)
        if not d:
            continue
        delta_days = (d - today).days
        if delta_days == 0:
            continue

        month_delta = _month_index(d) - _month_index(today)
        if month_delta == 0:
            months_away = 1
        else:
            months_away = min(abs(month_delta) + 1, 6)

        if delta_days < 0:
            palette = _PAST_GREENS
        else:
            palette = _FUTURE_REDS

        idx = min(months_away - 1, len(palette) - 1)
        color = palette[idx]
        fill = PatternFill(fill_type="solid", fgColor=color.replace("#", ""))
        use_white_text = months_away >= 5

        for c in range(1, ws.max_column + 1):
            target = ws.cell(row=row, column=c)
            target.fill = fill
            if use_white_text:
                target.font = target.font.copy(color="FFFFFF")


def _excel_color_by_column_palette(
    ws, df: pd.DataFrame, col: str, data_start_row: int = 2
):
    if col not in df.columns:
        return
    values = df[col].astype(str).str.strip().tolist()
    uniques = []
    for v in values:
        if v and v not in uniques:
            uniques.append(v)
    if not uniques:
        return

    palette = [
        "F8BBD0",
        "E1BEE7",
        "C5CAE9",
        "BBDEFB",
        "B2EBF2",
        "B2DFDB",
        "C8E6C9",
        "DCEDC8",
        "FFF9C4",
        "FFE0B2",
    ]
    color_map = {u: palette[i % len(palette)] for i, u in enumerate(uniques)}

    for row_idx, v in enumerate(values, start=data_start_row):
        color = color_map.get(v)
        if not color:
            continue
        fill = PatternFill(fill_type="solid", fgColor=color)
        for c in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=c).fill = fill


def _excel_insert_title_row(ws, title: str, blank_row: bool = True) -> int:
    insert_rows = 2 if blank_row else 1
    ws.insert_rows(1, amount=insert_rows)
    last_col = ws.max_column
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    cell = ws.cell(row=1, column=1)
    cell.value = title
    cell.font = Font(size=20, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    if blank_row:
        ws.row_dimensions[2].height = 6
    return 3 if blank_row else 2


def _excel_set_freeze_panes(ws, header_row: int):
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


def _excel_autofit_columns(
    ws, header_row: int, min_width: int = 8, max_width: int = 40
):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row in ws.iter_rows(
            min_row=header_row,
            max_row=ws.max_row,
            min_col=col_idx,
            max_col=col_idx,
        ):
            cell = row[0]
            if cell.value is None:
                continue
            val = str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _excel_set_right_align_cols(ws, header_row: int, col_names: List[str]):
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val is None:
            continue
        header_map[str(val).strip().upper()] = col_idx

    for name in col_names:
        col_idx = header_map.get(str(name).strip().upper())
        if not col_idx:
            continue
        for row in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is None:
                continue
            if cell.alignment:
                cell.alignment = cell.alignment.copy(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="right")


def _excel_set_cell_border(cell, left=None, right=None, top=None, bottom=None):
    border = cell.border
    cell.border = Border(
        left=left or border.left,
        right=right or border.right,
        top=top or border.top,
        bottom=bottom or border.bottom,
        diagonal=border.diagonal,
        diagonalUp=border.diagonalUp,
        diagonalDown=border.diagonalDown,
        outline=border.outline,
        vertical=border.vertical,
        horizontal=border.horizontal,
    )


def _excel_apply_thick_outline(
    ws, min_row: int, max_row: int, min_col: int, max_col: int
):
    thick = Side(border_style="thick", color="000000")
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            left = thick if col == min_col else None
            right = thick if col == max_col else None
            top = thick if row == min_row else None
            bottom = thick if row == max_row else None
            if not (left or right or top or bottom):
                continue
            _excel_set_cell_border(
                ws.cell(row=row, column=col),
                left=left,
                right=right,
                top=top,
                bottom=bottom,
            )


def _excel_outline_date_blocks(
    ws,
    header_row: int,
    model_col: str = "MODELO",
    date_col: str = "FECHA",
    talla_col: str = "TALLA",
):
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val is None:
            continue
        header_map[str(val).strip().upper()] = col_idx
    model_idx = header_map.get(model_col.upper())
    date_idx = header_map.get(date_col.upper())
    talla_idx = header_map.get(talla_col.upper())
    if not model_idx or not date_idx:
        return

    data_start = header_row + 1
    blocks: List[Tuple[int, int]] = []
    cur_model = None
    cur_date = None
    block_start = None

    for row in range(data_start, ws.max_row + 1):
        model = ws.cell(row=row, column=model_idx).value
        fecha = ws.cell(row=row, column=date_idx).value
        talla_val = ws.cell(row=row, column=talla_idx).value if talla_idx else ""
        talla_txt = str(talla_val).strip().upper() if talla_val else ""

        if talla_txt in ("TOTAL MODELO", "TOTAL GENERAL"):
            if block_start is not None:
                blocks.append((block_start, row - 1))
                block_start = None
            cur_model = None
            cur_date = None
            if talla_txt == "TOTAL MODELO":
                _excel_apply_thick_outline(ws, row, row, 1, ws.max_column)
            continue

        if not model or not fecha:
            if block_start is not None:
                blocks.append((block_start, row - 1))
                block_start = None
            cur_model = None
            cur_date = None
            continue

        model_txt = str(model).strip()
        date_txt = str(fecha).strip()
        if cur_model is None:
            cur_model = model_txt
            cur_date = date_txt
            block_start = row
        elif model_txt == cur_model and date_txt == cur_date:
            continue
        else:
            if block_start is not None:
                blocks.append((block_start, row - 1))
            cur_model = model_txt
            cur_date = date_txt
            block_start = row

    if block_start is not None:
        blocks.append((block_start, ws.max_row))

    for start, end in blocks:
        if start <= end:
            _excel_apply_thick_outline(ws, start, end, 1, ws.max_column)


def _excel_outline_model_blocks(ws, header_row: int, model_col: str = "MODELO"):
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val is None:
            continue
        header_map[str(val).strip().upper()] = col_idx
    model_idx = header_map.get(model_col.upper())
    if not model_idx:
        return

    data_start = header_row + 1
    blocks: List[Tuple[int, int]] = []
    cur_model = None
    block_start = None

    for row in range(data_start, ws.max_row + 1):
        model = ws.cell(row=row, column=model_idx).value
        if not model:
            if block_start is not None:
                blocks.append((block_start, row - 1))
                block_start = None
            cur_model = None
            continue
        model_txt = str(model).strip()
        if cur_model is None:
            cur_model = model_txt
            block_start = row
        elif model_txt == cur_model:
            continue
        else:
            if block_start is not None:
                blocks.append((block_start, row - 1))
            cur_model = model_txt
            block_start = row

    if block_start is not None:
        blocks.append((block_start, ws.max_row))

    for start, end in blocks:
        if start <= end:
            _excel_apply_thick_outline(ws, start, end, 1, ws.max_column)


# -----------------------
# Ops: status + listados
# -----------------------
def op_status(args):
    paths = {
        "inv": args.inv,
        "prev": args.prev,
        "talleres": args.talleres,
        "clientes": args.clientes,
        "export_dir": args.export_dir or "",
        "backup_dir": args.backup_dir or "",
    }
    exists = {
        k: (Path(v).exists() if v else False)
        for k, v in paths.items()
        if k in ("inv", "prev", "talleres", "clientes")
    }
    mgr = _make_mgr(args)
    modelos = sorted(
        set(
            list(getattr(mgr.inventory, "almacen", {}).keys())
            + list(getattr(mgr.inventory, "info_modelos", {}).keys())
        )
    )
    return _ok(paths=paths, exists=exists, num_modelos=len(modelos))


def op_preview_stock(args):
    mgr = _make_mgr(args)
    modelo_sel = (args.modelo or "").strip().upper()
    talla_sel = norm_talla(args.talla or "")

    rows = []
    almacen = getattr(mgr.inventory, "almacen", {}) or {}
    info_modelos = getattr(mgr.inventory, "info_modelos", {}) or {}

    modelos = sorted(almacen.keys())
    for m in modelos:
        if modelo_sel and m != modelo_sel:
            continue
        for t, q in sorted((almacen.get(m, {}) or {}).items(), key=lambda x: str(x[0])):
            if talla_sel and norm_talla(t) != talla_sel:
                continue
            info = info_modelos.get(m, {}) or {}
            rows.append(
                {
                    "MODELO": m,
                    "DESCRIPCION": info.get("descripcion", ""),
                    "COLOR": info.get("color", ""),
                    "CLIENTE": info.get("cliente", ""),
                    "TALLA": t,
                    "STOCK": q,
                }
            )

    cols = ["MODELO", "DESCRIPCION", "COLOR", "CLIENTE", "TALLA", "STOCK"]
    return _ok(columns=cols, rows=rows)


def op_list_modelos(args):
    mgr = _make_mgr(args)

    # universo de modelos: inventario + info_modelos (por si hay info sin stock)
    almacen = getattr(mgr.inventory, "almacen", {}) or {}
    info_modelos = getattr(mgr.inventory, "info_modelos", {}) or {}

    modelos = sorted(set(list(almacen.keys()) + list(info_modelos.keys())))

    items = []
    for m in modelos:
        info = info_modelos.get(m, {}) or {}
        items.append(
            {
                "modelo": m,
                "descripcion": info.get("descripcion", "") or "",
                "color": info.get("color", "") or "",
                "cliente": info.get("cliente", "") or "",
            }
        )
    return _ok(items=items)


def op_list_tallas(args):
    """
    Devuelve tallas “conocidas” para un modelo, igual que en Streamlit:
    - Stock (almacen)
    - Pedidos pendientes
    - Órdenes de fabricación

    Esto hace que el selector de talla en Movimientos sea útil aunque el modelo
    no tenga stock todavía (pero sí pedidos/fabricación).
    """
    mgr = _make_mgr(args)
    modelo = (args.modelo or "").strip().upper()
    if not modelo:
        return _ok(items=[])

    tallas = set()

    # 1) Stock real
    almacen = getattr(mgr.inventory, "almacen", {}) or {}
    tallas_map = almacen.get(modelo, {}) or {}
    for t in tallas_map.keys():
        tt = norm_talla(t)
        if tt:
            tallas.add(tt)

    # 2) Pendientes
    try:
        for _, p in mgr.prevision.list_pendings():
            if str(p.get("modelo", "")).strip().upper() == modelo:
                tt = norm_talla(p.get("talla", ""))
                if tt:
                    tallas.add(tt)
    except Exception:
        pass

    # 3) Fabricación
    try:
        for _, f in mgr.prevision.list_fabrication():
            if str(f.get("modelo", "")).strip().upper() == modelo:
                tt = norm_talla(f.get("talla", ""))
                if tt:
                    tallas.add(tt)
    except Exception:
        pass

    return _ok(items=sorted(tallas))


def op_list_pendings(args):
    mgr = _make_mgr(args)
    pend = mgr.prevision.list_pendings()
    rows = []
    for idx, p in pend:
        info = mgr.inventory.info_modelos.get(
            p["modelo"], mgr.prevision.info_modelos.get(p["modelo"], {})
        )
        rows.append(
            {
                "IDX": idx,
                "MODELO": p.get("modelo", ""),
                "DESCRIPCION": (info or {}).get("descripcion", ""),
                "COLOR": (info or {}).get("color", ""),
                "TALLA": p.get("talla", ""),
                "CANTIDAD": p.get("cantidad", 0),
                "PEDIDO": p.get("pedido", ""),
                "NUMERO_PEDIDO": p.get("numero_pedido", ""),
                "CLIENTE": p.get("cliente", ""),
                "FECHA": p.get("fecha", ""),
            }
        )
    cols = [
        "IDX",
        "MODELO",
        "DESCRIPCION",
        "COLOR",
        "TALLA",
        "CANTIDAD",
        "PEDIDO",
        "NUMERO_PEDIDO",
        "CLIENTE",
        "FECHA",
    ]
    return _ok(columns=cols, rows=rows)


def op_list_fabrication(args):
    mgr = _make_mgr(args)
    items = mgr.prevision.list_fabrication()
    rows = []
    for idx, it in items:
        info = mgr.inventory.info_modelos.get(
            it["modelo"], mgr.prevision.info_modelos.get(it["modelo"], {})
        )
        rows.append(
            {
                "IDX": idx,
                "MODELO": it.get("modelo", ""),
                "DESCRIPCION": (info or {}).get("descripcion", ""),
                "COLOR": (info or {}).get("color", ""),
                "TALLA": it.get("talla", ""),
                "CANTIDAD": it.get("cantidad", 0),
                "FECHA": it.get("fecha", ""),
            }
        )
    cols = ["IDX", "MODELO", "DESCRIPCION", "COLOR", "TALLA", "CANTIDAD", "FECHA"]
    return _ok(columns=cols, rows=rows)


def op_calc_estimated(args):
    mgr = _make_mgr(args)
    est = mgr.prevision.calc_estimated_stock(mgr.inventory)
    df = pd.DataFrame(est)
    if not df.empty:
        df = df.sort_values(["modelo", "talla"])
        rows = df.to_dict(orient="records")
        cols = list(df.columns)
    else:
        rows, cols = [], ["modelo", "talla", "stock_estimado"]
    return _ok(columns=cols, rows=rows)


# -----------------------
# Ops: movimientos
# -----------------------
def op_register_entry(args):
    mgr = _make_mgr(args)
    modelo = (args.modelo or "").strip().upper()
    talla = norm_talla(args.talla or "")
    cantidad = int(args.cantidad or 0)
    if not modelo or not talla or cantidad <= 0:
        return _fail("BAD_INPUT", "modelo+talla+cantidad>0 obligatorios")

    def _do():
        return mgr.inventory.register_entry(
            modelo=modelo,
            talla=talla,
            cantidad=cantidad,
            taller=(args.taller or "") or "",
            fecha=(args.fecha or "") or None,
            proveedor=(args.proveedor or "") or "",
            observaciones=(args.obs or "") or "",
        )

    _, log, err = _capture_io(_do)

    # JSON limpio + logs opcionales (muy útil para debug)
    return _ok(message="ENTRY_OK", log=log.strip(), stderr=err.strip())


def op_register_exit(args):
    mgr = _make_mgr(args)
    modelo = (args.modelo or "").strip().upper()
    talla = norm_talla(args.talla or "")
    cantidad = int(args.cantidad or 0)
    pedido = norm_codigo(args.pedido or "")
    albaran = norm_codigo(args.albaran or "")
    if not modelo or not talla or cantidad <= 0 or not pedido or not albaran:
        return _fail("BAD_INPUT", "modelo+talla+cantidad>0+pedido+albaran obligatorios")

    def _do():
        return mgr.inventory.register_exit(
            modelo=modelo,
            talla=talla,
            cantidad=cantidad,
            cliente=(args.cliente or "") or "",
            pedido=pedido,
            albaran=albaran,
            fecha=(args.fecha or "") or None,
        )

    ok, log, err = _capture_io(_do)
    return _ok(
        message="EXIT_OK", registered=bool(ok), log=log.strip(), stderr=err.strip()
    )


# -----------------------
# Ops: pendientes CRUD
# -----------------------
def op_add_pending(args):
    mgr = _make_mgr(args)
    modelo = (args.modelo or "").strip().upper()
    talla = norm_talla(args.talla or "")
    cantidad = int(args.cantidad or 0)
    pedido = norm_codigo(args.pedido or "")
    if not modelo or not talla or cantidad <= 0 or not pedido:
        return _fail("BAD_INPUT", "modelo+talla+cantidad>0+pedido obligatorios")

    mgr.prevision.register_pending(
        modelo=modelo,
        talla=talla,
        cantidad=cantidad,
        pedido=pedido,
        cliente=(args.cliente or "") or "",
        fecha=(args.fecha or "") or None,
        numero_pedido=norm_codigo(args.numero_pedido or "") or None,
    )
    return _ok(message="PENDING_ADDED")


def op_edit_pending(args):
    mgr = _make_mgr(args)
    idx = int(args.idx or 0)
    if idx <= 0:
        return _fail("BAD_INPUT", "idx>0 obligatorio")

    # vacíos => None (sin cambio)
    def _none_if_empty(s: Optional[str]):
        if s is None:
            return None
        ss = str(s).strip()
        return ss if ss else None

    modelo = _none_if_empty(args.modelo)
    if modelo:
        modelo = modelo.upper()

    talla = _none_if_empty(args.talla)
    if talla:
        talla = norm_talla(talla)

    pedido = _none_if_empty(args.pedido)
    if pedido:
        pedido = norm_codigo(pedido)

    num = _none_if_empty(args.numero_pedido)
    if num:
        num = norm_codigo(num)

    cantidad = None
    if args.cantidad is not None and str(args.cantidad).strip() != "":
        try:
            cantidad = int(args.cantidad)
        except Exception:
            return _fail("BAD_INPUT", "cantidad debe ser int")

    mgr.prevision.edit_pending(
        idx,
        modelo=modelo,
        talla=talla,
        cantidad=cantidad,
        pedido=pedido,
        cliente=_none_if_empty(args.cliente),
        fecha=_none_if_empty(args.fecha),
        numero_pedido=num,
    )
    return _ok(message="PENDING_EDITED")


def op_delete_pending(args):
    mgr = _make_mgr(args)
    idx = int(args.idx or 0)
    if idx <= 0:
        return _fail("BAD_INPUT", "idx>0 obligatorio")
    mgr.prevision.delete_pending(idx)
    return _ok(message="PENDING_DELETED")


# -----------------------
# Ops: fabricación CRUD
# -----------------------
def op_add_fabrication(args):
    mgr = _make_mgr(args)
    modelo = (args.modelo or "").strip().upper()
    talla = norm_talla(args.talla or "")
    cantidad = int(args.cantidad or 0)
    if not modelo or not talla or cantidad <= 0:
        return _fail("BAD_INPUT", "modelo+talla+cantidad>0 obligatorios")

    mgr.prevision.register_order(
        modelo, talla, cantidad, fecha=(args.fecha or "") or None
    )
    return _ok(message="FAB_ADDED")


def op_edit_fabrication_qty(args):
    mgr = _make_mgr(args)
    idx = int(args.idx or 0)
    qty = int(args.cantidad or 0)
    if idx <= 0:
        return _fail("BAD_INPUT", "idx>0 obligatorio")
    mgr.prevision.edit_fabrication_qty(idx, qty)
    return _ok(message="FAB_EDITED")


# -----------------------
# Ops: auditoría (preview + aplicar + regularizar)
# -----------------------
def op_audit_preview(args):
    mgr = _make_mgr(args)
    solo_modelo = (args.modelo or "").strip().upper() or None
    cambios = mgr.inventory.audit_and_fix_stock(aplicar=False, solo_modelo=solo_modelo)
    return _ok(columns=list(cambios[0].keys()) if cambios else [], rows=cambios or [])


def op_audit_apply(args):
    mgr = _make_mgr(args)
    payload = json.loads(args.payload_json or "{}")
    cambios = payload.get("cambios", [])
    if not isinstance(cambios, list) or not cambios:
        return _fail("BAD_INPUT", "payload.cambios debe ser lista no vacía")
    n = mgr.inventory.apply_stock_fixes(cambios)
    return _ok(message="AUDIT_APPLIED", updated=int(n))


def op_audit_regularize(args):
    mgr = _make_mgr(args)
    payload = json.loads(args.payload_json or "{}")
    cambios = payload.get("cambios", [])
    fecha = payload.get("fecha") or datetime.now().strftime("%Y-%m-%d")
    obs = payload.get("obs") or "Ajuste auditoría (GUI)"
    if not isinstance(cambios, list) or not cambios:
        return _fail("BAD_INPUT", "payload.cambios debe ser lista no vacía")
    n = mgr.inventory.regularize_history_to_current(
        cambios, fecha=fecha, observacion=obs
    )
    return _ok(message="AUDIT_REGULARIZED", created=int(n))


# -----------------------
# Ops: saneos
# -----------------------
def op_fix_negatives_to_zero(args):
    mgr = _make_mgr(args)
    cambios = []
    for modelo, tallas in list(mgr.inventory.almacen.items()):
        for talla, val in list(tallas.items()):
            try:
                v = int(val)
            except Exception:
                try:
                    v = int(float(str(val).replace(",", ".")))
                except Exception:
                    v = 0
            if v < 0:
                cambios.append(
                    {"MODELO": modelo, "TALLA": talla, "ANTES": v, "AJUSTADO_A": 0}
                )
                mgr.inventory.almacen[modelo][talla] = 0

    ruta_log = ""
    if cambios:
        mgr.inventory.save()
        export_dir = (
            Path(args.export_dir)
            if args.export_dir
            else Path(
                getattr(
                    mgr,
                    "EXPORT_DIR",
                    Path(mgr.ds_inventario.path).parent / "EXPORTAR_CSV",
                )
            )
        )
        export_dir.mkdir(parents=True, exist_ok=True)
        ruta_log = str(export_dir / f"ajuste_negativos_{_timestamp()}.csv")
        pd.DataFrame(cambios).to_csv(ruta_log, index=False, encoding="utf-8-sig")

    return _ok(
        message="FIX_NEGATIVES_DONE",
        changed=len(cambios),
        log_path=ruta_log,
        rows=cambios,
    )


def op_fix_bad_stock_values(args):
    mgr = _make_mgr(args)
    import math

    cambios = []
    for modelo, tallas in list(mgr.inventory.almacen.items()):
        for talla, val in list(tallas.items()):
            original_val = val
            original_talla = talla
            try:
                if val is None:
                    nuevo_val = 0
                elif isinstance(val, float) and math.isnan(val):
                    nuevo_val = 0
                elif isinstance(val, int):
                    nuevo_val = int(val)
                elif isinstance(val, float):
                    nuevo_val = int(val)
                elif isinstance(val, str):
                    s = val.strip().lower()
                    if s in ("nan", "none", ""):
                        nuevo_val = 0
                    else:
                        s = s.replace(",", ".")
                        nuevo_val = int(float(s))
                else:
                    nuevo_val = 0
            except Exception:
                nuevo_val = 0

            talla_str = str(talla).strip().upper()
            bad_key = (
                talla is None
                or (isinstance(talla, float) and math.isnan(talla))
                or talla_str in ("", "NAN", "NA", "NULL")
            )
            motivo = "VALOR_INVALIDO"
            if bad_key and nuevo_val != 0:
                nuevo_val = 0
                motivo = "TALLA_ANOMALA->VALOR_0"

            if nuevo_val != original_val:
                mgr.inventory.almacen[modelo][talla] = nuevo_val
                cambios.append(
                    {
                        "MODELO": modelo,
                        "TALLA": original_talla,
                        "ANTES": original_val,
                        "AJUSTADO_A": nuevo_val,
                        "MOTIVO": motivo,
                    }
                )

    if cambios:
        mgr.inventory.save()

    return _ok(message="FIX_BAD_VALUES_DONE", changed=len(cambios), rows=cambios)


def op_purge_bad_talla_keys(args):
    mgr = _make_mgr(args)
    import math

    only_zero = bool(int(args.only_zero or 1))
    bad_labels = {"", "NAN", "NA", "NULL"}
    borradas = []
    for modelo, tallas in list(mgr.inventory.almacen.items()):
        for talla in list(tallas.keys()):
            talla_str = "" if talla is None else str(talla).strip().upper()
            is_bad = (
                talla is None
                or (isinstance(talla, float) and math.isnan(talla))
                or talla_str in bad_labels
            )
            if not is_bad:
                continue

            val = tallas.get(talla, 0)
            try:
                v = int(val)
            except Exception:
                try:
                    v = int(float(str(val).replace(",", ".")))
                except Exception:
                    v = 0

            if (only_zero and v == 0) or (not only_zero):
                borradas.append({"MODELO": modelo, "TALLA": talla, "VALOR": v})
                del mgr.inventory.almacen[modelo][talla]

    ruta_log = ""
    if borradas:
        mgr.inventory.save()
        export_dir = (
            Path(args.export_dir)
            if args.export_dir
            else Path(
                getattr(
                    mgr,
                    "EXPORT_DIR",
                    Path(mgr.ds_inventario.path).parent / "EXPORTAR_CSV",
                )
            )
        )
        export_dir.mkdir(parents=True, exist_ok=True)
        ruta_log = str(export_dir / f"purga_tallas_anomalas_{_timestamp()}.csv")
        pd.DataFrame(borradas).to_csv(ruta_log, index=False, encoding="utf-8-sig")

    return _ok(
        message="PURGE_BAD_TALLAS_DONE",
        deleted=len(borradas),
        log_path=ruta_log,
        rows=borradas,
    )


# -----------------------
# Ops: catálogo/maestros
# -----------------------
def op_list_catalog(args):
    mgr = _make_mgr(args)
    modelos = []
    for m, info in sorted(mgr.inventory.info_modelos.items()):
        modelos.append({"MODELO": m, **(info or {})})
    talleres = [
        {"NOMBRE": t.nombre, "CONTACTO": t.contacto or ""}
        for t in mgr.workshops.list_all()
    ]
    clientes = [
        {"NOMBRE": c.nombre, "CONTACTO": c.contacto or ""}
        for c in mgr.clients.list_all()
    ]
    return _ok(modelos=modelos, talleres=talleres, clientes=clientes)


def op_update_model_info(args):
    mgr = _make_mgr(args)
    modelo = (args.modelo or "").strip().upper()
    if not modelo:
        return _fail("BAD_INPUT", "modelo obligatorio")
    mgr.inventory.update_model_info(
        modelo=modelo,
        descripcion=(args.descripcion or None),
        color=(args.color or None),
        cliente=(args.cliente or None),
    )
    return _ok(message="MODEL_INFO_UPDATED")


def op_add_taller(args):
    mgr = _make_mgr(args)
    nombre = (args.nombre or "").strip()
    if not nombre:
        return _fail("BAD_INPUT", "nombre obligatorio")
    mgr.workshops.add(nombre, (args.contacto or None))
    return _ok(message="TALLER_ADDED")


def op_add_cliente(args):
    mgr = _make_mgr(args)
    nombre = (args.nombre or "").strip()
    if not nombre:
        return _fail("BAD_INPUT", "nombre obligatorio")
    mgr.clients.add(nombre, (args.contacto or None))
    return _ok(message="CLIENTE_ADDED")


# -----------------------
# Ops: importaciones (Excel)
# - usamos ruta de archivo ya guardado por Next (args.excel_path)
# - o ruta fija desde gestor (ALBARANES_EXCEL / PEDIDOS_EXCEL)
# -----------------------
def _procesar_albaranes_df(
    mgr: GestorStock, df: pd.DataFrame, modo: str, simular: bool
) -> Dict[str, Any]:
    columnas = [
        "CodigoArticulo",
        "DesTalla",
        "Total",
        "SuPedido",
        "FechaAlbaran",
        "NumeroAlbaran",
    ]
    if not all(col in df.columns for col in columnas):
        raise ValueError(f"Faltan columnas necesarias: {columnas}")

    ya_registrado = defaultdict(int)
    for s in mgr.inventory.historial_salidas:
        try:
            k = (
                str(s.get("modelo", "")).strip().upper(),
                norm_talla(s.get("talla", "")),
                norm_codigo(s.get("pedido", "")),
                norm_codigo(s.get("albaran", "")),
            )
            ya_registrado[k] += int(s.get("cantidad", 0) or 0)
        except Exception:
            continue

    lineas = []
    for _, fila in df.iterrows():
        modelo = str(fila["CodigoArticulo"]).strip().upper()
        talla = norm_talla(fila["DesTalla"])
        val_total = fila["Total"]
        cantidad_excel = (
            int(val_total)
            if not pd.isna(val_total) and str(val_total).strip() != ""
            else 0
        )
        val_pedido = fila["SuPedido"]
        pedido = norm_codigo("" if pd.isna(val_pedido) else val_pedido)
        val_albar = fila["NumeroAlbaran"]
        albaran = norm_codigo("" if pd.isna(val_albar) else val_albar)
        val_fecha = fila["FechaAlbaran"]
        fecha = parse_fecha_excel(None if pd.isna(val_fecha) else val_fecha)
        k = (modelo, talla, pedido, albaran)
        qty_prev = ya_registrado.get(k, 0)

        lineas.append(
            {
                "modelo": modelo,
                "talla": talla,
                "pedido": pedido,
                "albaran": albaran,
                "fecha": fecha,
                "cantidad_excel": cantidad_excel,
                "ya_prev": qty_prev,
            }
        )

    nuevas_salidas = 0
    import_rows = []

    pedidos_antes = list(mgr.prevision.pedidos)

    for L in lineas:
        modelo = L["modelo"]
        talla = L["talla"]
        pedido = L["pedido"]
        albaran = L["albaran"]
        fecha = L["fecha"]
        qty_excel = int(L["cantidad_excel"])
        qty_prev = int(L["ya_prev"])

        if modo == "d" and qty_prev > 0:
            aplicar = max(qty_excel - qty_prev, 0)
        elif modo == "i" and qty_prev > 0:
            aplicar = 0
        else:
            aplicar = qty_excel

        if aplicar <= 0:
            continue

        if not simular:
            cliente = ""
            for p in mgr.prevision.pedidos:
                if (
                    str(p.get("modelo", "")).strip().upper() == modelo
                    and norm_talla(p.get("talla", "")) == talla
                    and p.get("pedido", "") == pedido
                ):
                    cliente = p.get("cliente", "") or ""
                    if cliente:
                        break
            if not cliente:
                cliente = (
                    mgr.prevision.info_modelos.get(modelo, {}).get("cliente", "") or ""
                )
            mgr.inventory.register_exit(
                modelo=modelo,
                talla=talla,
                cantidad=aplicar,
                cliente=cliente,
                pedido=pedido,
                albaran=albaran,
                fecha=fecha,
            )

        nuevas_salidas += aplicar
        import_rows.append(
            {
                "FECHA": fecha,
                "MODELO": modelo,
                "TALLA": talla,
                "CANTIDAD": aplicar,
                "PEDIDO": pedido,
                "ALBARAN": albaran,
                "CLIENTE": "",
            }
        )

    pedidos_despues = list(mgr.prevision.pedidos)
    set_antes = {
        (p["modelo"], norm_talla(p["talla"]), p["pedido"]) for p in pedidos_antes
    }
    set_despues = {
        (p["modelo"], norm_talla(p["talla"]), p["pedido"]) for p in pedidos_despues
    }
    servidos = set_antes - set_despues

    return {
        "nuevas_salidas": nuevas_salidas,
        "import_rows": import_rows,
        "servidos": [
            {"MODELO": m, "TALLA": t, "PEDIDO": ped} for (m, t, ped) in servidos
        ],
    }


def _procesar_pedidos_df(
    mgr: GestorStock, df: pd.DataFrame, simular: bool
) -> Dict[str, Any]:
    columnas = [
        "CodigoArticulo",
        "DesTalla",
        "UnidadesPendientes",
        "SuPedido",
        "FechaEntrega",
        "NumeroPedido",
    ]
    if not all(col in df.columns for col in columnas):
        raise ValueError(f"Faltan columnas necesarias: {columnas}")

    ya = {
        (
            str(p.get("modelo", "")).strip().upper(),
            norm_talla(p.get("talla", "")),
            p.get("pedido", ""),
        )
        for p in mgr.prevision.pedidos
    }

    nuevos, duplicados = 0, 0
    import_rows = []

    for _, fila in df.iterrows():
        modelo = str(fila["CodigoArticulo"]).strip().upper()
        talla = norm_talla(fila["DesTalla"])
        val = fila["UnidadesPendientes"]
        cantidad = int(val) if not pd.isna(val) else 0
        pedido = norm_codigo(fila["SuPedido"])
        numero_pedido = norm_codigo(fila["NumeroPedido"])
        fecha = parse_fecha_excel(fila["FechaEntrega"])
        cliente_resuelto = (
            mgr.inventory.info_modelos.get(modelo, {}).get("cliente", "") or ""
        )

        k = (modelo, talla, pedido)
        if k in ya:
            duplicados += 1
            continue

        if not simular:
            mgr.prevision.register_pending(
                modelo=modelo,
                talla=talla,
                cantidad=int(cantidad),
                pedido=pedido,
                cliente=cliente_resuelto,
                fecha=fecha or None,
                numero_pedido=numero_pedido or None,
            )

        nuevos += 1
        import_rows.append(
            {
                "FECHA": fecha,
                "PEDIDO": pedido,
                "NUMERO_PEDIDO": numero_pedido,
                "MODELO": modelo,
                "TALLA": talla,
                "CANTIDAD": int(cantidad),
                "CLIENTE": cliente_resuelto,
            }
        )

    return {"nuevos": nuevos, "duplicados": duplicados, "import_rows": import_rows}


def op_import_albaranes(args):
    mgr = _make_mgr(args)
    modo = (args.modo or "d").strip().lower()
    simular = bool(int(args.simular or 0))
    skip = int(args.skip or 25)

    # excel_path tiene prioridad
    ruta = args.excel_path or getattr(mgr, "ALBARANES_EXCEL", None)
    if not ruta:
        return _fail("MISSING_PATH", "excel_path o ALBARANES_EXCEL requerido")
    df = pd.read_excel(ruta, skiprows=skip)
    out = _procesar_albaranes_df(mgr, df, modo=modo, simular=simular)
    return _ok(message="IMPORT_ALBARANES_OK", **out)


def op_import_pedidos(args):
    mgr = _make_mgr(args)
    simular = bool(int(args.simular or 0))
    skip = int(args.skip or 26)

    ruta = args.excel_path or getattr(mgr, "PEDIDOS_EXCEL", None)
    if not ruta:
        return _fail("MISSING_PATH", "excel_path o PEDIDOS_EXCEL requerido")
    df = pd.read_excel(ruta, skiprows=skip)
    out = _procesar_pedidos_df(mgr, df, simular=simular)
    return _ok(message="IMPORT_PEDIDOS_OK", **out)


# -----------------------
# Ops: backups
# -----------------------
def op_backup_create(args):
    mgr = _make_mgr(args)
    base_dir = (
        Path(args.backup_dir)
        if args.backup_dir
        else Path(
            getattr(mgr, "BACKUP_DIR", Path(mgr.ds_inventario.path).parent / "backups")
        )
    )
    base_dir.mkdir(parents=True, exist_ok=True)

    fecha = _timestamp()
    ruta_datos = base_dir / f"datos_almacen_{fecha}.json"
    ruta_prevision = base_dir / f"prevision_{fecha}.json"

    shutil.copyfile(mgr.ds_inventario.path, ruta_datos)
    shutil.copyfile(mgr.ds_prevision.path, ruta_prevision)

    return _ok(message="BACKUP_CREATED", files=[str(ruta_datos), str(ruta_prevision)])


def op_backup_list(args):
    mgr = _make_mgr(args)
    base_dir = (
        Path(args.backup_dir)
        if args.backup_dir
        else Path(
            getattr(mgr, "BACKUP_DIR", Path(mgr.ds_inventario.path).parent / "backups")
        )
    )
    base_dir.mkdir(parents=True, exist_ok=True)
    files = sorted([p.name for p in base_dir.glob("*.json")], reverse=True)
    return _ok(message="BACKUP_LIST", files=files, dir=str(base_dir))


def op_backup_restore(args):
    mgr = _make_mgr(args)
    base_dir = (
        Path(args.backup_dir)
        if args.backup_dir
        else Path(
            getattr(mgr, "BACKUP_DIR", Path(mgr.ds_inventario.path).parent / "backups")
        )
    )
    base_dir.mkdir(parents=True, exist_ok=True)

    name = (args.name or "").strip()
    if not name:
        return _fail("BAD_INPUT", "name obligatorio")
    origen = base_dir / name
    if not origen.exists():
        return _fail("NOT_FOUND", f"no existe {origen}")

    if "datos_almacen" in name:
        destino = Path(mgr.ds_inventario.path)
    elif "prevision" in name:
        destino = Path(mgr.ds_prevision.path)
    else:
        return _fail("BAD_INPUT", "backup debe incluir 'datos_almacen' o 'prevision'")

    shutil.copyfile(origen, destino)
    return _ok(message="BACKUP_RESTORED", restored=name, dest=str(destino))


# -----------------------
# Ops: exports (ZIP)
# -----------------------
def _zip_dir(source_dir: Path, out_zip: Path):
    with zipfile.ZipFile(out_zip, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for p in source_dir.rglob("*"):
            if p.is_file():
                z.write(p, arcname=p.relative_to(source_dir))


def op_export_csv_pack(args):
    mgr = _make_mgr(args)
    export_dir = (
        Path(args.export_dir)
        if args.export_dir
        else Path(
            getattr(
                mgr, "EXPORT_DIR", Path(mgr.ds_inventario.path).parent / "EXPORTAR_CSV"
            )
        )
    )
    export_dir.mkdir(parents=True, exist_ok=True)

    if hasattr(mgr, "_exportar_todos_los_datos"):
        mgr._exportar_todos_los_datos()
    else:
        return _fail("MISSING_BACKEND", "mgr._exportar_todos_los_datos no existe")

    if not args.out:
        return _fail("MISSING_OUT", "--out requerido")
    out_zip = Path(args.out)
    out_zip.parent.mkdir(parents=True, exist_ok=True)
    _zip_dir(export_dir, out_zip)
    return _ok(
        message="EXPORT_CSV_PACK_OK", out=str(out_zip), export_dir=str(export_dir)
    )


def op_export_stock_negativo(args):
    mgr = _make_mgr(args)
    export_dir = (
        Path(args.export_dir)
        if args.export_dir
        else Path(
            getattr(
                mgr, "EXPORT_DIR", Path(mgr.ds_inventario.path).parent / "EXPORTAR_CSV"
            )
        )
    )
    export_dir.mkdir(parents=True, exist_ok=True)

    if hasattr(mgr, "_exportar_stock_negativo"):
        mgr._exportar_stock_negativo()
    else:
        return _fail("MISSING_BACKEND", "mgr._exportar_stock_negativo no existe")

    if not args.out:
        return _fail("MISSING_OUT", "--out requerido")
    out_zip = Path(args.out)
    out_zip.parent.mkdir(parents=True, exist_ok=True)
    _zip_dir(export_dir, out_zip)
    return _ok(
        message="EXPORT_STOCK_NEG_OK", out=str(out_zip), export_dir=str(export_dir)
    )


def op_export_excel_pack(args):
    mgr = _make_mgr(args)
    export_dir = (
        Path(args.export_dir)
        if args.export_dir
        else Path(
            getattr(
                mgr, "EXPORT_DIR", Path(mgr.ds_inventario.path).parent / "EXPORTAR_CSV"
            )
        )
    )
    export_dir.mkdir(parents=True, exist_ok=True)

    if hasattr(mgr, "_exportar_todos_los_datos"):
        mgr._exportar_todos_los_datos()
    else:
        return _fail("MISSING_BACKEND", "mgr._exportar_todos_los_datos no existe")

    hoy_str = datetime.now().strftime("%Y-%m-%d")
    hoy_title = datetime.now().strftime("%d-%m-%Y")

    config = {
        "00": ("stock_actual", ["STOCK", "stock"], "stock"),
        "03": ("pedidos_pendientes", None, None),
        "04": ("ordenes_fabricacion", None, None),
        "05": ("stock_estimado", ["STOCK_ESTIMADO", "stock_estimado"], "estimado"),
        "06": ("orden_corte_sugerida", None, None),
    }

    generados: List[str] = []

    title_map = {
        "00": "01 - STOCK ACTUAL",
        "03": "02 - PEDIDOS PENDIENTES",
        "04": "03 - ORDENES FABRICACION",
        "05": "04 - STOCK ESTIMADO",
        "06": "05 - ORDEN CORTE SUGERIDA",
    }

    for prefijo, (base_name, qty_candidates, _) in config.items():
        patrones = [
            f"{prefijo}_{base_name}_*.csv",
            f"*{prefijo}*{base_name}*.csv",
        ]
        csv_path: Optional[Path] = None
        for pat in patrones:
            matches = sorted(export_dir.glob(pat))
            if matches:
                csv_path = matches[-1]
                break

        if not csv_path or not csv_path.exists():
            continue

        df = pd.read_csv(csv_path, sep=";", dtype=str)
        xlsx_name = f"IMPRIMIR_{prefijo}_{base_name}_{hoy_str}.xlsx"
        xlsx_path = export_dir / xlsx_name

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Hoja1")
            ws = writer.sheets["Hoja1"]

            title = f"{title_map.get(prefijo, base_name.upper())} - {hoy_title}"
            header_row = _excel_insert_title_row(ws, title, blank_row=True)
            data_start = header_row + 1

            if prefijo in ("00", "05"):
                _excel_color_stock_ranges(
                    ws, df, qty_col_candidates=qty_candidates, data_start_row=data_start
                )
            elif prefijo == "03":
                _excel_color_pend_by_month(
                    ws, df, date_col="FECHA", data_start_row=data_start
                )
            elif prefijo == "04":
                _excel_color_by_column_palette(
                    ws, df, col="FECHA", data_start_row=data_start
                )
            elif prefijo == "06":
                _excel_color_by_column_palette(
                    ws, df, col="MODELO", data_start_row=data_start
                )

            if prefijo in ("00", "03", "04", "05"):
                _excel_highlight_totals_by_talla(
                    ws, df, talla_col="TALLA", data_start_row=data_start
                )

            highlight_last = prefijo in ("00", "03", "04", "05")
            _excel_yellow_header_and_total(
                ws, header_row=header_row, highlight_last=highlight_last
            )
            _excel_add_borders(ws, min_row=header_row, min_col=1)

            if prefijo in ("03", "04"):
                _excel_outline_date_blocks(
                    ws,
                    header_row=header_row,
                    model_col="MODELO",
                    date_col="FECHA",
                    talla_col="TALLA",
                )
            if prefijo == "06":
                _excel_outline_model_blocks(ws, header_row=header_row, model_col="MODELO")

            if prefijo == "00":
                _excel_set_right_align_cols(ws, header_row, ["TALLA", "STOCK"])
            elif prefijo == "03":
                _excel_set_right_align_cols(ws, header_row, ["TALLA", "CANTIDAD"])
            elif prefijo == "04":
                _excel_set_right_align_cols(ws, header_row, ["TALLA", "CANTIDAD"])
            elif prefijo == "05":
                _excel_set_right_align_cols(
                    ws, header_row, ["TALLA", "STOCK_ESTIMADO"]
                )

            _excel_autofit_columns(ws, header_row)
            _excel_set_freeze_panes(ws, header_row)

        generados.append(str(xlsx_path))

    if not args.out:
        return _fail("MISSING_OUT", "--out requerido")
    out_zip = Path(args.out)
    out_zip.parent.mkdir(parents=True, exist_ok=True)

    _zip_dir(export_dir, out_zip)
    return _ok(
        message="EXPORT_EXCEL_PACK_OK",
        out=str(out_zip),
        generated=generados,
        export_dir=str(export_dir),
    )


# -----------------------
# Main dispatcher
# -----------------------
OPS = {
    # status + listados
    "status": op_status,
    "preview_stock": op_preview_stock,
    "list_pendings": op_list_pendings,
    "list_fabrication": op_list_fabrication,
    "calc_estimated": op_calc_estimated,
    # movimientos
    "register_entry": op_register_entry,
    "register_exit": op_register_exit,
    # pendientes
    "add_pending": op_add_pending,
    "edit_pending": op_edit_pending,
    "delete_pending": op_delete_pending,
    # fabricación
    "add_fabrication": op_add_fabrication,
    "edit_fabrication_qty": op_edit_fabrication_qty,
    # auditoría
    "audit_preview": op_audit_preview,
    "audit_apply": op_audit_apply,
    "audit_regularize": op_audit_regularize,
    # saneos
    "fix_negatives_to_zero": op_fix_negatives_to_zero,
    "fix_bad_stock_values": op_fix_bad_stock_values,
    "purge_bad_talla_keys": op_purge_bad_talla_keys,
    # catálogo/maestros
    "list_catalog": op_list_catalog,
    "update_model_info": op_update_model_info,
    "add_taller": op_add_taller,
    "add_cliente": op_add_cliente,
    # importaciones
    "import_albaranes": op_import_albaranes,
    "import_pedidos": op_import_pedidos,
    # backups
    "backup_create": op_backup_create,
    "backup_list": op_backup_list,
    "backup_restore": op_backup_restore,
    # exports (zip)
    "export_csv_pack": op_export_csv_pack,
    "export_stock_negativo": op_export_stock_negativo,
    "export_excel_pack": op_export_excel_pack,
    "list_modelos": op_list_modelos,
    "list_tallas": op_list_tallas,
}


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser()
    p.add_argument("--op", required=True)

    # paths json
    p.add_argument(
        "--inv", default=_read_env_path("GLOBALIA_INV_PATH", "datos_almacen.json")
    )
    p.add_argument(
        "--prev", default=_read_env_path("GLOBALIA_PREV_PATH", "prevision.json")
    )
    p.add_argument(
        "--talleres", default=_read_env_path("GLOBALIA_TALLERES_PATH", "talleres.json")
    )
    p.add_argument(
        "--clientes", default=_read_env_path("GLOBALIA_CLIENTES_PATH", "clientes.json")
    )

    # dirs
    p.add_argument(
        "--export-dir",
        dest="export_dir",
        default=_read_env_path("GLOBALIA_EXPORT_DIR", ""),
    )
    p.add_argument(
        "--backup-dir",
        dest="backup_dir",
        default=_read_env_path("GLOBALIA_BACKUP_DIR", ""),
    )

    # out zip
    p.add_argument("--out", default="")

    # gen params
    p.add_argument("--modelo", default="")
    p.add_argument("--talla", default="")
    p.add_argument("--cantidad", default=None)
    p.add_argument("--fecha", default="")
    p.add_argument("--cliente", default="")
    p.add_argument("--taller", default="")
    p.add_argument("--pedido", default="")
    p.add_argument("--albaran", default="")
    p.add_argument("--proveedor", default="")
    p.add_argument("--obs", default="")

    # pend/fab edits
    p.add_argument("--idx", default=None)
    p.add_argument("--numero-pedido", dest="numero_pedido", default="")
    p.add_argument("--payload-json", dest="payload_json", default="")

    # saneos
    p.add_argument("--only-zero", dest="only_zero", default="1")

    # imports
    p.add_argument("--excel-path", dest="excel_path", default="")
    p.add_argument("--modo", default="d")  # d/i/t
    p.add_argument("--simular", default="0")  # 0/1
    p.add_argument("--skip", default=None)

    # backups
    p.add_argument("--name", default="")

    # model info
    p.add_argument("--descripcion", default="")
    p.add_argument("--color", default="")
    p.add_argument("--nombre", default="")
    p.add_argument("--contacto", default="")

    return p


def main(argv=None) -> int:
    args = build_parser().parse_args(argv)
    op = args.op.strip()
    fn = OPS.get(op)
    if not fn:
        return _fail("UNKNOWN_OP", op)
    try:
        return int(fn(args) or 0)
    except Exception as e:
        return _fail("EXCEPTION", str(e))


if __name__ == "__main__":
    raise SystemExit(main())
