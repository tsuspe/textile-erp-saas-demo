#!/usr/bin/env python3
from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


REPORTS = {
    "00": {"name": "stock_actual", "align_right": ["TALLA", "STOCK"]},
    "03": {"name": "pedidos_pendientes", "align_right": ["TALLA", "CANTIDAD"]},
    "04": {"name": "ordenes_fabricacion", "align_right": ["TALLA", "CANTIDAD"]},
    "05": {"name": "stock_estimado", "align_right": ["TALLA", "STOCK_ESTIMADO"]},
    "06": {"name": "orden_corte_sugerida", "align_right": []},
}


def _latest_file(folder: Path, prefix: str) -> Optional[Path]:
    matches = sorted(folder.glob(f"IMPRIMIR_{prefix}_*.xlsx"))
    if not matches:
        return None
    return max(matches, key=lambda p: p.stat().st_mtime)


def _find_header_row(ws, header_name: str = "MODELO", max_scan: int = 5) -> Optional[int]:
    header_name = header_name.strip().upper()
    for row in range(1, max_scan + 1):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            if str(val).strip().upper() == header_name:
                return row
    return None


def _header_map(ws, header_row: int) -> Dict[str, int]:
    mapping = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            continue
        mapping[str(val).strip().upper()] = col
    return mapping


def _has_title_merge(ws) -> bool:
    for rng in ws.merged_cells.ranges:
        if rng.min_row == 1 and rng.max_row == 1 and rng.min_col == 1:
            return True
    return False


def _right_aligned_ratio(ws, header_row: int, col_idx: int) -> Tuple[int, int]:
    ok = 0
    total = 0
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is None:
            continue
        total += 1
        if cell.alignment and cell.alignment.horizontal == "right":
            ok += 1
    return ok, total


def _all_negative(ws, header_row: int, col_idx: int) -> bool:
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is None:
            continue
        try:
            val = float(str(cell.value).replace(",", "."))
        except Exception:
            continue
        if val >= 0:
            return False
    return True


def _has_thick_borders(ws) -> bool:
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            b = ws.cell(row=row, column=col).border
            for side in (b.left, b.right, b.top, b.bottom):
                if side and side.style == "thick":
                    return True
    return False


def _validate_file(path: Path, ref_path: Optional[Path] = None) -> List[str]:
    issues: List[str] = []
    wb = load_workbook(path)
    ws = wb.active

    header_row = _find_header_row(ws)
    if not header_row:
        issues.append("No se encontró la fila de cabecera (MODELO).")
        return issues

    if not _has_title_merge(ws):
        issues.append("No se detectó fila de título mergeada en la fila 1.")

    if ref_path and ref_path.exists():
        ref_wb = load_workbook(ref_path)
        ref_ws = ref_wb.active
        if ws.max_row != ref_ws.max_row:
            issues.append(
                f"Filas distintas vs referencia: {ws.max_row} vs {ref_ws.max_row}"
            )
        if ws.max_column != ref_ws.max_column:
            issues.append(
                f"Columnas distintas vs referencia: {ws.max_column} vs {ref_ws.max_column}"
            )

    mapping = _header_map(ws, header_row)
    for name in ("TALLA", "STOCK", "CANTIDAD", "STOCK_ESTIMADO"):
        if name in mapping:
            ok, total = _right_aligned_ratio(ws, header_row, mapping[name])
            if total and ok < total:
                issues.append(
                    f"Alineación derecha incompleta en {name}: {ok}/{total}."
                )

    if "STOCK_ESTIMADO" in mapping:
        if not _all_negative(ws, header_row, mapping["STOCK_ESTIMADO"]):
            issues.append("Hay valores no negativos en STOCK_ESTIMADO.")

    if not _has_thick_borders(ws):
        issues.append("No se detectaron bordes thick en la hoja.")

    return issues


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--export-dir", required=True)
    parser.add_argument("--out-zip", default="")
    parser.add_argument("--ref-dir", default="")
    parser.add_argument("--no-generate", action="store_true")
    parser.add_argument("--inv", default="")
    parser.add_argument("--prev", default="")
    parser.add_argument("--talleres", default="")
    parser.add_argument("--clientes", default="")
    parser.add_argument("--backup-dir", default="")
    args = parser.parse_args()

    export_dir = Path(args.export_dir)
    export_dir.mkdir(parents=True, exist_ok=True)
    out_zip = Path(args.out_zip) if args.out_zip else export_dir / "verify_export.zip"

    cli_path = Path(__file__).resolve().parent / "cli.py"
    if not args.no_generate:
        cmd = [
            sys.executable,
            str(cli_path),
            "--op",
            "export_excel_pack",
            "--out",
            str(out_zip),
            "--export-dir",
            str(export_dir),
        ]
        if args.inv:
            cmd += ["--inv", args.inv]
        if args.prev:
            cmd += ["--prev", args.prev]
        if args.talleres:
            cmd += ["--talleres", args.talleres]
        if args.clientes:
            cmd += ["--clientes", args.clientes]
        if args.backup_dir:
            cmd += ["--backup-dir", args.backup_dir]
        subprocess.run(cmd, check=True)

    ref_dir = Path(args.ref_dir) if args.ref_dir else None
    exit_code = 0
    for prefix, meta in REPORTS.items():
        out_file = _latest_file(export_dir, prefix)
        if not out_file:
            print(f"[WARN] No se encontró IMPRIMIR_{prefix}_*.xlsx en {export_dir}")
            exit_code = 1
            continue
        ref_file = _latest_file(ref_dir, prefix) if ref_dir else None
        issues = _validate_file(out_file, ref_file)
        if issues:
            exit_code = 1
            print(f"[FAIL] {out_file.name}")
            for it in issues:
                print(f"  - {it}")
        else:
            print(f"[OK] {out_file.name}")

    return exit_code


if __name__ == "__main__":
    raise SystemExit(main())
